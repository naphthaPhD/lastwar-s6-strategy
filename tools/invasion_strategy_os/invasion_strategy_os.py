from __future__ import annotations

import argparse
import csv
import inspect
import json
import re
import shutil
import sys
from dataclasses import asdict, dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any
from urllib.parse import quote
from urllib.request import urlopen
from zoneinfo import ZoneInfo

import networkx as nx
from pyvis.network import Network

try:
    from .simulation import build_invasion_simulation as build_state_invasion_simulation
    from .simulation import build_briefing_input
except ImportError:
    from simulation import build_invasion_simulation as build_state_invasion_simulation
    from simulation import build_briefing_input


JST = ZoneInfo("Asia/Tokyo")
CENTRAL_COORD_PREFIX = "\u4e2d\u592e"
DEFAULT_FISHERY_TYPE = "\u6f01\u5834"
DEFAULT_CITY_TYPE = "\u90fd\u5e02"
DEFAULT_ALTAR_TYPE = "\u796d\u58c7"
DEFAULT_CENTRAL_EMPTY_TYPE = "\u4e2d\u592e\u7a7a\u767d"
DEFAULT_ANCESTRAL_TEMPLE_TYPE = "\u7956\u970a\u795e\u6bbf"
CENTRAL_REFERENCE_EMPTY_COORDS = {(10, 10), (10, 11), (11, 10), (11, 11)}
SELF_AREA = "#534"
ALLY_AREAS = {"#509", "#440", "#511"}
ENEMY_AREAS = {"#503", "#480", "#523", "#476"}
SELF_SERVERS = {"534"}
ALLY_SERVERS = {"509", "440", "511"}
ENEMY_SERVERS = {"503", "480", "523", "476"}
DEFAULT_SELF_OWNERS = {"JDX"}
UNOWNED_OWNER_VALUES = {"", "unknown", "none", "neutral", "\u672a\u53d6\u5f97", "\u672a\u767b\u9332", "\u4e2d\u7acb", "\u4e2d\u7acb/\u672a\u767b\u9332"}
DESTROYED_KEYWORDS = ("\u7834\u58ca", "destroyed", "ruined")

NODE_ALIASES = {
    "id": ["id", "node_id", "key", "キー", "ID", "拠点ID", "位置キー"],
    "name": ["name", "拠点名", "名称", "施設名", "座標"],
    "type": ["type", "種別", "施設種別"],
    "owner": ["owner", "所有連盟", "連盟", "alliance", "現在連盟"],
    "protect_until": ["protect_until", "保護終了時刻", "保護終了", "保護明け", "protect", "保護切れ"],
    "x": ["x", "X", "座標X", "game_x"],
    "y": ["y", "Y", "座標Y", "game_y"],
    "importance": ["importance", "重要度", "priority", "優先度"],
    "adjacent": ["adjacent", "隣接", "接続情報", "neighbors", "neighbours"],
    "area": ["area", "エリア", "server", "サーバ"],
}

EDGE_ALIASES = {
    "from": ["from", "source", "from_id", "起点"],
    "to": ["to", "target", "to_id", "終点"],
    "weight": ["weight", "重み"],
}

PACT_ALIASES = {
    "alliance_a": ["連盟A", "alliance_a", "Alliance A", "A"],
    "alliance_b": ["連盟B", "alliance_b", "Alliance B", "B"],
    "starts_at": ["開始日", "start", "starts_at"],
    "ends_at": ["終了日", "end", "ends_at", "valid_until"],
    "status": ["状態", "status"],
    "safety_window": ["安全期間（現地時間）", "安全期間", "safety_window"],
    "confirmed_at": ["確認日", "confirmed_at", "confirmed"],
    "memo": ["メモ", "memo", "note"],
}


@dataclass(frozen=True)
class Node:
    id: str
    name: str
    type: str
    owner: str
    protect_until: str | None
    x: float
    y: float
    importance: float
    area: str = ""
    acquired_at: str | None = None
    status: str = ""
    memo: str = ""


@dataclass(frozen=True)
class Edge:
    source: str
    target: str
    weight: float = 1.0


@dataclass(frozen=True)
class AlliancePower:
    server: str
    alliance: str
    alliance_name: str
    power: int | None
    side: str = ""
    overall_rank: int | None = None
    server_rank: int | None = None


def read_config(path: str | None) -> dict[str, Any]:
    if not path:
        return {}
    with Path(path).open("r", encoding="utf-8") as fh:
        return json.load(fh)


def read_csv_rows(source: dict[str, Any], base_dir: Path) -> list[dict[str, str]]:
    source_type = source.get("type", "csv")
    if source_type == "csv":
        raw_path = Path(source["path"])
        path = raw_path if raw_path.is_absolute() else base_dir / raw_path
        text = path.read_text(encoding="utf-8-sig")
    elif source_type == "google_sheet_csv":
        url = google_sheet_csv_url(source)
        with urlopen(url, timeout=30) as response:
            text = response.read().decode("utf-8-sig")
    elif source_type == "url_csv":
        with urlopen(source["url"], timeout=30) as response:
            text = response.read().decode("utf-8-sig")
    else:
        raise ValueError(f"Unsupported source type: {source_type}")

    if source.get("format") == "management_table":
        return parse_management_table_rows(text)
    return list(csv.DictReader(text.splitlines()))


def parse_management_table_rows(text: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row_number, row in enumerate(csv.reader(text.splitlines()), start=1):
        padded = row + [""] * max(0, 20 - len(row))
        coord = padded[1].strip() or padded[0].strip()
        area = padded[18].strip()
        position_key = padded[19].strip()
        if row_number == 1 or coord in {"座標", "座標(自動)"}:
            continue
        if not coord or not area or not position_key:
            continue
        x, y = local_coord_xy(coord)
        rows.append(
            {
                "id": position_key,
                "name": coord,
                "type": padded[2].strip(),
                "owner": padded[3].strip(),
                "protect_until": padded[5].strip(),
                "x": str(x),
                "y": str(y),
                "area": area,
                "acquired_at": padded[4].strip(),
                "status": padded[7].strip(),
                "memo": padded[11].strip(),
            }
        )
    return rows


def normalize_alliance_tag(value: str | None) -> str:
    return re.sub(r"\s+", "", str(value or "").strip())


def parse_pact_datetime(value: str | None, tz: ZoneInfo, end_of_day: bool = False) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y/%m/%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S%z"):
        try:
            parsed = datetime.strptime(text, fmt)
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=tz)
            return parsed
        except ValueError:
            continue
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            parsed_date = datetime.strptime(text, fmt).date()
            day_time = time(23, 59, 59) if end_of_day else time(0, 0, 0)
            return datetime.combine(parsed_date, day_time, tzinfo=tz)
        except ValueError:
            continue
    try:
        parsed = datetime.fromisoformat(text)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=tz)
        return parsed
    except ValueError:
        return None


def pact_confidence(confirmed_at: datetime | None, now: datetime) -> str:
    if confirmed_at is None:
        return "unknown"
    days = (now - confirmed_at).total_seconds() / 86400
    if days <= 2:
        return "high"
    if days <= 7:
        return "medium"
    return "low"


def load_pacts(rows: list[dict[str, str]], now: datetime, tz: ZoneInfo, source: dict[str, Any]) -> dict[str, Any]:
    active_records: dict[tuple[str, str], dict[str, Any]] = {}
    inactive_markers = ("期限切れ", "無効", "未締結", "解除", "expired", "inactive")

    for row_number, row in enumerate(rows, start=2):
        alliance_a = normalize_alliance_tag(pick(row, PACT_ALIASES, "alliance_a"))
        alliance_b = normalize_alliance_tag(pick(row, PACT_ALIASES, "alliance_b"))
        if not alliance_a or not alliance_b or alliance_a == alliance_b:
            continue

        status = pick(row, PACT_ALIASES, "status", "有効") or "有効"
        status_key = status.strip().lower()
        if any(marker in status_key for marker in inactive_markers):
            continue
        if status and "有効" not in status and status_key not in {"active", "valid"}:
            continue

        starts_at = parse_pact_datetime(pick(row, PACT_ALIASES, "starts_at"), tz)
        ends_at = parse_pact_datetime(pick(row, PACT_ALIASES, "ends_at"), tz, end_of_day=True)
        confirmed_at = parse_pact_datetime(pick(row, PACT_ALIASES, "confirmed_at"), tz)
        if starts_at and starts_at > now:
            continue
        if ends_at and ends_at < now:
            continue

        pair_key = tuple(sorted((alliance_a, alliance_b), key=lambda value: value.lower()))
        record = {
            "alliance_a": alliance_a,
            "alliance_b": alliance_b,
            "pair": list(pair_key),
            "status": status,
            "starts_at": starts_at.isoformat() if starts_at else None,
            "ends_at": ends_at.isoformat() if ends_at else None,
            "confirmed_at": confirmed_at.isoformat() if confirmed_at else None,
            "confidence": pact_confidence(confirmed_at, now),
            "safety_window": pick(row, PACT_ALIASES, "safety_window"),
            "memo": pick(row, PACT_ALIASES, "memo"),
            "source_row": row_number,
        }
        previous = active_records.get(pair_key)
        previous_confirmed = parse_pact_datetime(previous.get("confirmed_at") if previous else None, tz) if previous else None
        if previous is None or (confirmed_at or datetime.min.replace(tzinfo=tz)) >= (previous_confirmed or datetime.min.replace(tzinfo=tz)):
            active_records[pair_key] = record

    records = sorted(active_records.values(), key=lambda item: (item["pair"][0], item["pair"][1]))
    partners: dict[str, set[str]] = {}
    for record in records:
        a = str(record["alliance_a"])
        b = str(record["alliance_b"])
        partners.setdefault(a, set()).add(b)
        partners.setdefault(b, set()).add(a)

    confidence_summary: dict[str, int] = {}
    for record in records:
        confidence = str(record.get("confidence") or "unknown")
        confidence_summary[confidence] = confidence_summary.get(confidence, 0) + 1

    return {
        "generated_at": now.isoformat(),
        "source": {
            "type": source.get("type"),
            "spreadsheet_id": source.get("spreadsheet_id"),
            "gid": source.get("gid"),
            "sheet_name": source.get("sheet_name"),
        },
        "active": records,
        "partners_by_alliance": {owner: sorted(values) for owner, values in sorted(partners.items())},
        "active_count": len(records),
        "confidence_summary": confidence_summary,
    }


def local_coord_xy(coord: str) -> tuple[float, float]:
    central_match = re.match(r"^中央-(\d+)-(\d+)$", coord.strip())
    if central_match:
        row_text, col_text = central_match.groups()
        row = int(row_text)
        col = int(col_text)
        if row < 1 or row > 20 or col < 1 or col > 20:
            return 0.0, 0.0
        return float(24 + (col - 1) * 50), float(1024 - row * 50)

    match = re.match(r"^([A-Ka-j])-(\d+)$", coord.strip())
    if not match:
        return 0.0, 0.0
    letter, number_text = match.groups()
    number = int(number_text)
    if letter.isupper():
        axis = [20, 99, 199, 299, 399, 499, 599, 699, 799, 899, 978]
        row_index = ord(letter) - ord("A")
        col_index = (number - 1) // 2
        if row_index < 0 or row_index >= len(axis) or col_index < 0 or col_index >= len(axis):
            return 0.0, 0.0
        return float(axis[col_index]), float(axis[row_index])
    row_index = ord(letter) - ord("a")
    col_index = (number - 2) // 2
    if row_index < 0 or row_index >= 10 or col_index < 0 or col_index >= 10:
        return 0.0, 0.0
    return float(49 + col_index * 100), float(49 + row_index * 100)


def google_sheet_csv_url(source: dict[str, Any]) -> str:
    spreadsheet_id = source["spreadsheet_id"]
    if source.get("gid") is not None:
        gid = source["gid"]
        return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid={gid}"
    if source.get("sheet_name"):
        sheet_name = quote(str(source["sheet_name"]))
        return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/gviz/tq?tqx=out:csv&sheet={sheet_name}"
    raise ValueError("google_sheet_csv source needs either gid or sheet_name")


def pick(row: dict[str, str], aliases: dict[str, list[str]], field: str, default: str = "") -> str:
    for key in aliases[field]:
        if key in row and row[key] is not None:
            return str(row[key]).strip()
    return default


def parse_float(value: str, default: float = 0.0) -> float:
    if value == "":
        return default
    return float(str(value).replace(",", ""))


def parse_coord_letter(value: str) -> str:
    match = re.search(r"(?:^|:)([A-Za-z])-\d+$", value.strip())
    return match.group(1) if match else ""


def is_reference_central_fishery(row: int, col: int) -> bool:
    if row < 1 or row > 20 or col < 1 or col > 20:
        return False
    if row in {1, 2, 3, 18, 19, 20}:
        return True
    if row in {4, 17}:
        return col <= 4 or col >= 17
    return col <= 3 or col >= 18


def central_reference_type(coord: str, config: dict[str, Any]) -> str | None:
    if not config.get("central_type_from_reference"):
        return None
    match = re.match(rf"^{CENTRAL_COORD_PREFIX}-(\d+)-(\d+)$", coord.strip())
    if not match:
        return None
    row, col = (int(match.group(1)), int(match.group(2)))
    if (row, col) in CENTRAL_REFERENCE_EMPTY_COORDS:
        return str(config.get("central_empty_type", DEFAULT_CENTRAL_EMPTY_TYPE))
    if is_reference_central_fishery(row, col):
        return str(config.get("fishery_type", DEFAULT_FISHERY_TYPE))
    return str(config.get("altar_type", DEFAULT_ALTAR_TYPE))


def normalize_node_type(raw_type: str, coord: str, config: dict[str, Any]) -> str:
    node_type = raw_type or "unknown"
    reference_type = central_reference_type(coord, config)
    if reference_type:
        return reference_type
    if not config.get("type_from_coord_case"):
        return node_type
    isolated_types = {str(value) for value in config.get("isolated_types", [])}
    if node_type in isolated_types:
        return node_type
    letter = parse_coord_letter(coord)
    if not letter:
        return node_type
    if letter.islower():
        return str(config.get("city_type", "city"))
    if letter.isupper():
        return str(config.get("fishery_type", "fishery"))
    return node_type


def apply_area_offset(x: float, y: float, area: str, config: dict[str, Any]) -> tuple[float, float]:
    offsets = config.get("area_offsets", {})
    offset = offsets.get(area, [0, 0])
    if len(offset) != 2:
        raise ValueError(f"area_offsets for {area} must have exactly two numbers")
    return x + float(offset[0]), y + float(offset[1])


def add_central_ancestral_temple_node(
    nodes: dict[str, Node],
    config: dict[str, Any],
    importance_by_type: dict[str, float],
    area_filter: set[str],
    excluded_areas: set[str],
) -> None:
    if not config.get("central_ancestral_temple_node"):
        return
    area = str(config.get("central_area", CENTRAL_COORD_PREFIX))
    if area_filter and area not in area_filter:
        return
    if area in excluded_areas:
        return

    local_xy = config.get("central_ancestral_temple_local_xy", [499, 499])
    if len(local_xy) != 2:
        raise ValueError("central_ancestral_temple_local_xy must have exactly two numbers")
    node_type = str(config.get("ancestral_temple_type", DEFAULT_ANCESTRAL_TEMPLE_TYPE))
    name = str(config.get("central_ancestral_temple_name", node_type))
    node_id = str(config.get("central_ancestral_temple_id", f"{area}:{name}"))
    x, y = apply_area_offset(float(local_xy[0]), float(local_xy[1]), area, config)
    nodes[node_id] = Node(
        id=node_id,
        name=name,
        type=node_type,
        owner=str(config.get("central_ancestral_temple_owner", "unknown")),
        protect_until=None,
        x=x,
        y=y,
        importance=importance_by_type.get(node_type, float(config.get("central_ancestral_temple_importance", 10))),
        area=area,
        acquired_at=None,
        status=str(config.get("central_ancestral_temple_status", "")),
        memo="Synthetic 2x2 central node replacing Central-10-10, 10-11, 11-10, and 11-11.",
    )


def parse_optional_time(value: str | None, tz: ZoneInfo) -> datetime | None:
    if not value:
        return None
    normalized = value.strip()
    if not normalized:
        return None
    normalized = normalized.replace("Z", "+00:00")
    for fmt in ("%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S"):
        try:
            parsed = datetime.strptime(normalized, fmt)
            return parsed.replace(tzinfo=tz)
        except ValueError:
            pass
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=tz)
    return parsed.astimezone(tz)


def has_destroyed_marker(*values: str | None) -> bool:
    text = " ".join(str(value or "") for value in values).lower()
    return any(keyword.lower() in text for keyword in DESTROYED_KEYWORDS)


def is_destroyed_node(node: Node) -> bool:
    return node.type == DEFAULT_CITY_TYPE and has_destroyed_marker(node.status, node.memo, node.owner)


def is_destroyed_node_data(node_data: dict[str, Any]) -> bool:
    return str(node_data.get("type", "")) == DEFAULT_CITY_TYPE and has_destroyed_marker(
        str(node_data.get("status", "")),
        str(node_data.get("memo", "")),
        str(node_data.get("owner", "")),
    )


def split_neighbors(value: str) -> list[str]:
    return [part for part in re.split(r"[;,\s]+", value.strip()) if part]


def load_nodes(rows: list[dict[str, str]], config: dict[str, Any]) -> tuple[dict[str, Node], list[Edge]]:
    nodes: dict[str, Node] = {}
    adjacency_edges: list[Edge] = []
    area_filter = set(config.get("areas", []))
    excluded_areas = set(config.get("exclude_areas", []))
    excluded_types = {str(value) for value in config.get("exclude_types", [])}
    importance_by_type = {str(key): float(value) for key, value in config.get("importance_by_type", {}).items()}
    for row_number, row in enumerate(rows, start=2):
        node_id = pick(row, NODE_ALIASES, "id")
        if not node_id:
            raise ValueError(f"Node row {row_number} is missing id")
        area = pick(row, NODE_ALIASES, "area")
        if area_filter and area not in area_filter:
            continue
        if area in excluded_areas:
            continue
        node_name = pick(row, NODE_ALIASES, "name", node_id) or node_id
        node_type = normalize_node_type(pick(row, NODE_ALIASES, "type", "unknown"), node_name, config)
        if node_type in excluded_types:
            continue
        local_x = parse_float(pick(row, NODE_ALIASES, "x"))
        local_y = parse_float(pick(row, NODE_ALIASES, "y"))
        x, y = apply_area_offset(local_x, local_y, area, config)
        importance = parse_float(
            pick(row, NODE_ALIASES, "importance"),
            importance_by_type.get(node_type, 1.0),
        )
        node = Node(
            id=node_id,
            name=node_name,
            type=node_type,
            owner=pick(row, NODE_ALIASES, "owner", "unknown") or "unknown",
            protect_until=pick(row, NODE_ALIASES, "protect_until") or None,
            x=x,
            y=y,
            importance=importance,
            area=area,
            acquired_at=row.get("acquired_at") or None,
            status=row.get("status", ""),
            memo=row.get("memo", ""),
        )
        nodes[node.id] = node
        for neighbor_id in split_neighbors(pick(row, NODE_ALIASES, "adjacent")):
            adjacency_edges.append(Edge(node.id, neighbor_id, 1.0))
    add_central_ancestral_temple_node(nodes, config, importance_by_type, area_filter, excluded_areas)
    return nodes, adjacency_edges


def load_edges(rows: list[dict[str, str]]) -> list[Edge]:
    edges: list[Edge] = []
    for row_number, row in enumerate(rows, start=2):
        source = pick(row, EDGE_ALIASES, "from")
        target = pick(row, EDGE_ALIASES, "to")
        if not source and not target:
            continue
        if not source or not target:
            raise ValueError(f"Edge row {row_number} needs both from and to")
        weight = parse_float(pick(row, EDGE_ALIASES, "weight"), 1.0)
        edges.append(Edge(source, target, weight))
    return edges


def normalized_type_pair(source_type: str, target_type: str) -> tuple[str, str]:
    return tuple(sorted((source_type, target_type)))


def disallowed_edge_type_pairs(config: dict[str, Any]) -> set[tuple[str, str]]:
    configured = config.get("disallowed_edge_type_pairs", [[DEFAULT_CITY_TYPE, DEFAULT_CITY_TYPE]])
    return {
        normalized_type_pair(str(pair[0]), str(pair[1]))
        for pair in configured
        if isinstance(pair, list | tuple) and len(pair) == 2
    }


def filter_edges_by_type_rules(edges: list[Edge], nodes: dict[str, Node], config: dict[str, Any]) -> list[Edge]:
    isolated_types = {str(value) for value in config.get("isolated_types", [])}
    blocked_type_pairs = disallowed_edge_type_pairs(config)
    filtered: list[Edge] = []
    for edge in edges:
        source = nodes.get(edge.source)
        target = nodes.get(edge.target)
        if source is None or target is None:
            continue
        if is_destroyed_node(source) or is_destroyed_node(target):
            continue
        if source.type in isolated_types or target.type in isolated_types:
            continue
        if normalized_type_pair(source.type, target.type) in blocked_type_pairs:
            continue
        filtered.append(edge)
    return filtered


def derive_distance_edges(nodes: dict[str, Node], config: dict[str, Any]) -> list[Edge]:
    if config.get("type") == "coordinate":
        return derive_coordinate_edges(nodes, config)
    if config.get("type") != "distance":
        return []
    max_distance = float(config.get("max_distance", 80))
    same_area_only = bool(config.get("same_area_only", True))
    max_edges_per_node = int(config.get("max_edges_per_node", 6))
    isolated_types = {str(value) for value in config.get("isolated_types", [])}
    blocked_type_pairs = disallowed_edge_type_pairs(config)
    candidates: list[tuple[float, str, str]] = []
    node_items = sorted(nodes.items())
    for index, (source_id, source) in enumerate(node_items):
        if is_destroyed_node(source):
            continue
        if source.type in isolated_types:
            continue
        for target_id, target in node_items[index + 1 :]:
            if is_destroyed_node(target):
                continue
            if target.type in isolated_types:
                continue
            if normalized_type_pair(source.type, target.type) in blocked_type_pairs:
                continue
            if same_area_only and source.area != target.area:
                continue
            distance = ((source.x - target.x) ** 2 + (source.y - target.y) ** 2) ** 0.5
            if 0 < distance <= max_distance:
                candidates.append((distance, source_id, target_id))

    degree: dict[str, int] = {node_id: 0 for node_id in nodes}
    edges: list[Edge] = []
    for distance, source_id, target_id in sorted(candidates):
        if max_edges_per_node > 0 and (degree[source_id] >= max_edges_per_node or degree[target_id] >= max_edges_per_node):
            continue
        edges.append(Edge(source_id, target_id, round(distance, 3)))
        degree[source_id] += 1
        degree[target_id] += 1
    return edges


def parse_outer_fishery_coord(coord: str) -> tuple[int, int] | None:
    match = re.match(r"^([A-K])-(\d+)$", coord.strip())
    if not match:
        return None
    letter, number_text = match.groups()
    number = int(number_text)
    if number < 1 or number > 21 or number % 2 == 0:
        return None
    return ord(letter) - ord("A"), (number - 1) // 2


def parse_outer_city_coord(coord: str) -> tuple[int, int] | None:
    match = re.match(r"^([a-j])-(\d+)$", coord.strip())
    if not match:
        return None
    letter, number_text = match.groups()
    number = int(number_text)
    if number < 2 or number > 20 or number % 2 != 0:
        return None
    return ord(letter) - ord("a"), (number - 2) // 2


def parse_central_grid_coord(coord: str) -> tuple[int, int] | None:
    match = re.match(rf"^{CENTRAL_COORD_PREFIX}-(\d+)-(\d+)$", coord.strip())
    if not match:
        return None
    row, col = int(match.group(1)), int(match.group(2))
    if row < 1 or row > 20 or col < 1 or col > 20:
        return None
    return row, col


def edge_key(source_id: str, target_id: str) -> tuple[str, str]:
    return tuple(sorted((source_id, target_id)))


def boundary_variants(index: int, max_index: int) -> list[int]:
    if index <= 0 or index >= max_index:
        return [index]
    return [index - 1, index, index + 1]


def add_edge_once(edge_map: dict[tuple[str, str], Edge], source_id: str, target_id: str, weight: float = 1.0) -> None:
    if source_id == target_id:
        return
    key = edge_key(source_id, target_id)
    edge_map.setdefault(key, Edge(key[0], key[1], weight))


def facility_blocks_fishery_diagonal(
    facility_by_area_coord: dict[tuple[str, int, int], str],
    area: str,
    row: int,
    col: int,
    row_step: int,
    col_step: int,
) -> bool:
    if abs(row_step) != 1 or abs(col_step) != 1:
        return False
    city_row = row if row_step > 0 else row - 1
    city_col = col if col_step > 0 else col - 1
    return (area, city_row, city_col) in facility_by_area_coord


def outer_area_grid(nodes: dict[str, Node]) -> dict[str, tuple[int, int]]:
    areas: dict[str, list[Node]] = {}
    for node in nodes.values():
        if parse_outer_fishery_coord(node.name) is not None:
            areas.setdefault(node.area, []).append(node)
    grid: dict[str, tuple[int, int]] = {}
    for area, area_nodes in areas.items():
        min_x = min(node.x for node in area_nodes)
        min_y = min(node.y for node in area_nodes)
        grid[area] = (round((min_x - 20) / 1000), round((min_y - 20) / 1000))
    return grid


def derive_coordinate_edges(nodes: dict[str, Node], config: dict[str, Any]) -> list[Edge]:
    isolated_types = {str(value) for value in config.get("isolated_types", [])}
    fishery_type = str(config.get("fishery_type", DEFAULT_FISHERY_TYPE))
    city_type = str(config.get("city_type", DEFAULT_CITY_TYPE))
    trade_type = str(config.get("trade_type", "\u4ea4\u6613\u5730"))
    include_central = bool(config.get("include_central_fishery_edges", True))
    central_boundary_max_distance = float(config.get("central_boundary_max_distance", 76))

    outer_by_area_coord: dict[tuple[str, int, int], str] = {}
    outer_fishery_nodes: dict[str, Node] = {}
    outer_city_by_area_coord: dict[tuple[str, int, int], str] = {}
    outer_blocking_facility_by_area_coord: dict[tuple[str, int, int], str] = {}
    central_by_coord: dict[tuple[int, int], str] = {}
    central_fishery_nodes: dict[str, Node] = {}
    for node_id, node in nodes.items():
        if node.type == fishery_type:
            if node.type in isolated_types:
                continue
            outer_coord = parse_outer_fishery_coord(node.name)
            if outer_coord is not None:
                outer_by_area_coord[(node.area, outer_coord[0], outer_coord[1])] = node_id
                outer_fishery_nodes[node_id] = node
                continue
            central_coord = parse_central_grid_coord(node.name)
            if include_central and central_coord is not None:
                central_by_coord[central_coord] = node_id
                central_fishery_nodes[node_id] = node
            continue
        if node.type in {city_type, trade_type}:
            outer_city_coord = parse_outer_city_coord(node.name)
            if outer_city_coord is not None:
                outer_blocking_facility_by_area_coord[(node.area, outer_city_coord[0], outer_city_coord[1])] = node_id
                if node.type == city_type and not is_destroyed_node(node) and node.type not in isolated_types:
                    outer_city_by_area_coord[(node.area, outer_city_coord[0], outer_city_coord[1])] = node_id

    edges: dict[tuple[str, str], Edge] = {}

    grid_by_area = outer_area_grid(nodes)
    outer_by_global_coord: dict[tuple[int, int], str] = {}
    outer_global_meta: dict[tuple[int, int], tuple[str, int, int]] = {}
    for (area, row, col), source_id in sorted(outer_by_area_coord.items()):
        area_grid = grid_by_area.get(area)
        if area_grid is None:
            continue
        area_col, area_row = area_grid
        global_coord = (area_row * 11 + row, area_col * 11 + col)
        outer_by_global_coord[global_coord] = source_id
        outer_global_meta[global_coord] = (area, row, col)

    fishery_neighbor_steps = ((0, 1), (1, -1), (1, 0), (1, 1))
    for (row, col), source_id in sorted(outer_by_global_coord.items()):
        area, local_row, local_col = outer_global_meta[(row, col)]
        for row_step, col_step in fishery_neighbor_steps:
            if facility_blocks_fishery_diagonal(
                outer_blocking_facility_by_area_coord,
                area,
                local_row,
                local_col,
                row_step,
                col_step,
            ):
                continue
            target_id = outer_by_global_coord.get((row + row_step, col + col_step))
            if target_id:
                add_edge_once(edges, source_id, target_id)

    for (area, row, col), city_id in sorted(outer_city_by_area_coord.items()):
        for fishery_row, fishery_col in ((row, col), (row, col + 1), (row + 1, col), (row + 1, col + 1)):
            target_id = outer_by_area_coord.get((area, fishery_row, fishery_col))
            if target_id:
                add_edge_once(edges, city_id, target_id)

    for (row, col), source_id in sorted(central_by_coord.items()):
        for row_step, col_step in fishery_neighbor_steps:
            target_id = central_by_coord.get((row + row_step, col + col_step))
            if target_id:
                add_edge_once(edges, source_id, target_id)

    if central_boundary_max_distance > 0:
        for outer_id, outer_node in sorted(outer_fishery_nodes.items()):
            for central_id, central_node in sorted(central_fishery_nodes.items()):
                distance = ((outer_node.x - central_node.x) ** 2 + (outer_node.y - central_node.y) ** 2) ** 0.5
                if 0 < distance <= central_boundary_max_distance:
                    add_edge_once(edges, outer_id, central_id, round(distance, 3))

    return list(edges.values())


def build_graph(nodes: dict[str, Node], edges: list[Edge]) -> nx.Graph:
    graph = nx.Graph()
    for node in nodes.values():
        graph.add_node(node.id, **asdict(node))
    for edge in edges:
        if edge.source not in nodes or edge.target not in nodes:
            missing = [node_id for node_id in (edge.source, edge.target) if node_id not in nodes]
            raise ValueError(f"Edge references missing node(s): {', '.join(missing)}")
        graph.add_edge(edge.source, edge.target, weight=edge.weight)
    return graph


def is_major_node(graph: nx.Graph, node_id: str, config: dict[str, Any]) -> bool:
    node = graph.nodes[node_id]
    major_types = {str(value).lower() for value in config.get("major_types", ["city", "stronghold"])}
    major_importance = float(config.get("major_importance", 7))
    return str(node.get("type", "")).lower() in major_types or float(node.get("importance", 0)) >= major_importance


def analyze_graph(
    graph: nx.Graph,
    config: dict[str, Any],
    shortest_from: str | None,
    shortest_to: str | None,
) -> dict[str, Any]:
    components = [sorted(component) for component in nx.connected_components(graph)]
    articulation_points = sorted(nx.articulation_points(graph))
    degree_centrality = nx.degree_centrality(graph)
    betweenness_centrality = nx.betweenness_centrality(graph, weight="weight", normalized=True)
    isolated_nodes = sorted(nx.isolates(graph))

    shortest_path: list[str] | None = None
    if shortest_from and shortest_to:
        if shortest_from not in graph or shortest_to not in graph:
            shortest_path = None
        else:
            try:
                shortest_path = nx.shortest_path(graph, shortest_from, shortest_to, weight="weight")
            except nx.NetworkXNoPath:
                shortest_path = None

    critical_nodes = extract_chokes(graph, articulation_points, betweenness_centrality, config)

    return {
        "connected_components": components,
        "articulation_points": articulation_points,
        "critical_nodes": critical_nodes,
        "isolated_nodes": isolated_nodes,
        "shortest_path": {
            "from": shortest_from,
            "to": shortest_to,
            "path": shortest_path,
        },
        "degree_centrality": round_mapping(degree_centrality),
        "betweenness_centrality": round_mapping(betweenness_centrality),
    }


def extract_chokes(
    graph: nx.Graph,
    articulation_points: list[str],
    betweenness: dict[str, float],
    config: dict[str, Any],
) -> list[dict[str, Any]]:
    major_nodes = {node_id for node_id in graph.nodes if is_major_node(graph, node_id, config)}
    route_limited_degree = int(config.get("route_limited_degree", 2))
    chokes: list[dict[str, Any]] = []

    for index, node_id in enumerate(articulation_points, start=1):
        reduced = graph.copy()
        reduced.remove_node(node_id)
        split_components = [sorted(component) for component in nx.connected_components(reduced)]
        major_components = [sorted(set(component) & major_nodes) for component in split_components]
        major_components = [component for component in major_components if component]

        reasons = ["removal_disconnects_graph"]
        if len(major_components) >= 2:
            reasons.append("major_groups_split")
        if any(len(component) == 1 for component in major_components):
            reasons.append("major_node_isolated")
        if graph.degree(node_id) <= route_limited_degree:
            reasons.append("limited_route_degree")

        node = graph.nodes[node_id]
        score = (
            float(node.get("importance", 0))
            + graph.degree(node_id)
            + (betweenness.get(node_id, 0.0) * 10)
            + (3 if "major_groups_split" in reasons else 0)
            + (2 if "major_node_isolated" in reasons else 0)
        )
        chokes.append(
            {
                "choke_id": f"CHOKE-{index:02d}",
                "node_id": node_id,
                "name": node.get("name", node_id),
                "owner": node.get("owner", "unknown"),
                "type": node.get("type", "unknown"),
                "importance": node.get("importance", 0),
                "degree": graph.degree(node_id),
                "betweenness": round(betweenness.get(node_id, 0.0), 4),
                "score": round(score, 3),
                "reasons": reasons,
                "affected_major_groups": major_components,
                "components_after_removal": split_components,
            }
        )

    chokes.sort(key=lambda item: (-item["score"], item["node_id"]))
    for index, choke in enumerate(chokes, start=1):
        choke["choke_id"] = f"CHOKE-{index:02d}"
    return chokes


def round_mapping(values: dict[str, float]) -> dict[str, float]:
    return {key: round(value, 4) for key, value in sorted(values.items())}


def is_unowned_owner(owner: str | None) -> bool:
    normalized = str(owner or "").strip().lower()
    return normalized in UNOWNED_OWNER_VALUES


def display_owner(owner: str | None) -> str:
    return "\u672a\u53d6\u5f97" if is_unowned_owner(owner) else str(owner).strip()


def display_node_owner(node: Node) -> str:
    return "\u7834\u58ca" if is_destroyed_node(node) else display_owner(node.owner)


def area_affiliation(area: str) -> str:
    if area == SELF_AREA:
        return "self"
    if area in ALLY_AREAS:
        return "ally"
    if area in ENEMY_AREAS:
        return "enemy"
    return "neutral"


def server_affiliation(server: str) -> str:
    if server in SELF_SERVERS:
        return "self"
    if server in ALLY_SERVERS:
        return "ally"
    if server in ENEMY_SERVERS:
        return "enemy"
    return "enemy"


def owner_server_prefix(owner: str) -> str | None:
    match = re.match(r"^#?(\d{3})", owner.strip())
    return match.group(1) if match else None


def build_owner_affiliations(graph: nx.Graph, config: dict[str, Any]) -> dict[str, str]:
    self_owners = DEFAULT_SELF_OWNERS | {str(value).strip() for value in config.get("self_owners", [])}
    ally_owners = {str(value).strip() for value in config.get("ally_owners", [])}
    enemy_owners = {str(value).strip() for value in config.get("enemy_owners", [])}
    affiliations: dict[str, str] = {}
    for owner in self_owners:
        if owner:
            affiliations[owner] = "self"
    for owner in ally_owners:
        if owner:
            affiliations[owner] = "ally"
    for owner in enemy_owners:
        if owner:
            affiliations[owner] = "enemy"

    owner_area_counts: dict[str, dict[str, int]] = {}
    for _, node_data in graph.nodes(data=True):
        if is_destroyed_node_data(node_data):
            continue
        owner = str(node_data.get("owner", "")).strip()
        if is_unowned_owner(owner) or owner in affiliations:
            continue
        server = owner_server_prefix(owner)
        if server:
            affiliations[owner] = server_affiliation(server)
            continue
        affiliation = area_affiliation(str(node_data.get("area", "")))
        if affiliation == "neutral":
            continue
        counts = owner_area_counts.setdefault(owner, {"self": 0, "ally": 0, "enemy": 0})
        counts[affiliation] += 1

    for owner, counts in owner_area_counts.items():
        affiliations[owner] = sorted(
            counts.items(),
            key=lambda item: (-item[1], {"self": 0, "ally": 1, "enemy": 2}[item[0]]),
        )[0][0]
    return affiliations


def affiliation_color(affiliation: str) -> str:
    return {
        "self": "#2563eb",
        "ally": "#16a34a",
        "enemy": "#dc2626",
        "destroyed": "#6b7280",
    }.get(affiliation, "#f8fafc")


def strategic_affiliation(node: Node, owner_affiliations: dict[str, str]) -> str:
    if is_destroyed_node(node):
        return "destroyed"
    if is_unowned_owner(node.owner):
        return "unowned"
    return owner_affiliations.get(str(node.owner).strip(), area_affiliation(node.area))


def strategic_color(node: Node, owner_affiliations: dict[str, str] | None = None) -> str:
    if is_destroyed_node(node):
        return "#6b7280"
    if is_unowned_owner(node.owner):
        return "#f8fafc"
    if owner_affiliations is None:
        return affiliation_color(area_affiliation(node.area))
    return affiliation_color(strategic_affiliation(node, owner_affiliations))


def visual_node_size(node: Node, config: dict[str, Any]) -> float:
    node_min_size = float(config.get("node_min_size", 12))
    node_max_size = float(config.get("node_max_size", 55))
    node_size_multiplier = float(config.get("node_size_multiplier", 4))
    central_fishery_size = config.get("central_fishery_size")
    if (
        central_fishery_size is not None
        and node.area == CENTRAL_COORD_PREFIX
        and node.type == DEFAULT_FISHERY_TYPE
    ):
        return float(central_fishery_size)
    visual_size_by_type = {str(key): float(value) for key, value in config.get("visual_size_by_type", {}).items()}
    return visual_size_by_type.get(
        node.type,
        max(node_min_size, min(node_max_size, node_min_size + node.importance * node_size_multiplier)),
    )


def visual_font_size(node: Node, default_size: int, config: dict[str, Any]) -> int:
    if node.type == DEFAULT_FISHERY_TYPE:
        return int(config.get("fishery_font_size", max(9, default_size - 3)))
    return default_size


def visual_node_shape(node: Node, config: dict[str, Any]) -> str:
    visual_shape_by_type = {str(key): str(value) for key, value in config.get("visual_shape_by_type", {}).items()}
    return visual_shape_by_type.get(node.type, "dot")


def visual_position(node: Node, config: dict[str, Any]) -> tuple[float, float]:
    visual_gap = float(config.get("visual_area_gap", 0))
    tile_size = float(config.get("visual_area_tile_size", 1000))
    if visual_gap <= 0 or tile_size <= 0:
        return node.x, node.y
    column = int(node.x // tile_size)
    row = int(node.y // tile_size)
    return node.x + column * visual_gap, node.y + row * visual_gap


def protection_status(node: Node, now: datetime, warning_hours: float, tz: ZoneInfo) -> dict[str, Any]:
    protect_until = parse_optional_time(node.protect_until, tz)
    if protect_until is None:
        return {"status": "unknown", "hours_remaining": None, "protect_until": None}
    hours_remaining = (protect_until - now).total_seconds() / 3600
    if hours_remaining < 0:
        status = "expired"
    elif hours_remaining <= warning_hours:
        status = "soon"
    else:
        status = "protected"
    return {
        "status": status,
        "hours_remaining": round(hours_remaining, 2),
        "protect_until": protect_until.isoformat(),
    }


def parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r"[^\d]", "", str(value))
    return int(digits) if digits else None


def format_power(power: int | None) -> str:
    if power is None:
        return "-"
    if abs(power) >= 1_000_000_000:
        return f"{power / 1_000_000_000:.2f}B"
    if abs(power) >= 1_000_000:
        return f"{power / 1_000_000:.2f}M"
    return f"{power:,}"


def normalize_lookup_key(value: str | None) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def owner_lookup_keys(owner: str | None) -> list[str]:
    raw = str(owner or "").strip()
    keys: list[str] = []
    if raw:
        keys.append(normalize_lookup_key(raw))
    bracket_match = re.search(r"\[([^\]]+)\]", raw)
    if bracket_match:
        keys.append(normalize_lookup_key(bracket_match.group(1)))
    parts = raw.split(maxsplit=1)
    if raw.startswith("#") and len(parts) > 1:
        keys.append(normalize_lookup_key(parts[1]))
    return list(dict.fromkeys(key for key in keys if key))


def alliance_power_to_record(item: AlliancePower) -> dict[str, Any]:
    record = asdict(item)
    record["power_label"] = format_power(item.power)
    return record


def index_alliance_powers(records: list[AlliancePower]) -> dict[str, AlliancePower]:
    indexed: dict[str, AlliancePower] = {}
    for item in records:
        for key in owner_lookup_keys(item.alliance):
            indexed[key] = item
    return indexed


def read_alliance_power_cache(path: Path) -> dict[str, AlliancePower]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    records = [
        AlliancePower(
            server=str(item.get("server", "")),
            alliance=str(item.get("alliance", "")),
            alliance_name=str(item.get("alliance_name", "")),
            power=parse_int(item.get("power")),
            side=str(item.get("side", "")),
            overall_rank=parse_int(item.get("overall_rank")),
            server_rank=parse_int(item.get("server_rank")),
        )
        for item in payload.get("alliances", [])
    ]
    return index_alliance_powers(records)


def write_alliance_power_cache(path: Path, records: list[AlliancePower], source_path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source": str(source_path),
        "alliances": [alliance_power_to_record(item) for item in records],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_alliance_powers(config: dict[str, Any], base_dir: Path) -> dict[str, AlliancePower]:
    if not config:
        return {}

    cache_value = config.get("cache_json")
    cache_path = (base_dir / str(cache_value)) if cache_value else None
    source_type = str(config.get("type", "")).lower()
    if source_type not in {"xlsx", "excel"}:
        if cache_path and cache_path.exists():
            return read_alliance_power_cache(cache_path)
        return {}

    raw_path = Path(str(config.get("path", "")))
    source_path = raw_path if raw_path.is_absolute() else base_dir / raw_path
    if not source_path.exists():
        if cache_path and cache_path.exists():
            return read_alliance_power_cache(cache_path)
        return {}

    try:
        import openpyxl
    except ImportError as exc:
        if cache_path and cache_path.exists():
            return read_alliance_power_cache(cache_path)
        raise RuntimeError("openpyxl is required to read alliance power Excel data") from exc

    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    sheet_name = str(config.get("sheet", "Alliance Power"))
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Alliance power sheet not found: {sheet_name}")
    worksheet = workbook[sheet_name]
    header_row = int(config.get("header_row", 4))
    headers = [
        str(value).strip() if value is not None else ""
        for value in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    ]
    header_index = {name: index for index, name in enumerate(headers) if name}

    def cell(row: tuple[Any, ...], name: str) -> Any:
        index = header_index.get(name)
        if index is None or index >= len(row):
            return None
        return row[index]

    records: list[AlliancePower] = []
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        alliance = str(cell(row, "Alliance") or "").strip()
        if not alliance:
            continue
        records.append(
            AlliancePower(
                server=str(cell(row, "Server") or "").strip(),
                alliance=alliance,
                alliance_name=str(cell(row, "Alliance Name") or "").strip(),
                power=parse_int(cell(row, "Power")),
                side=str(cell(row, "Side") or "").strip(),
                overall_rank=parse_int(cell(row, "Overall Rank")),
                server_rank=parse_int(cell(row, "Server Rank")),
            )
        )

    if cache_path:
        write_alliance_power_cache(cache_path, records, source_path)
    return index_alliance_powers(records)


def alliance_power_for_owner(owner: str | None, powers: dict[str, AlliancePower]) -> AlliancePower | None:
    for key in owner_lookup_keys(owner):
        if key in powers:
            return powers[key]
    return None


def alliance_power_payload(owner: str | None, powers: dict[str, AlliancePower]) -> dict[str, Any] | None:
    item = alliance_power_for_owner(owner, powers)
    return alliance_power_to_record(item) if item else None


def write_html(
    graph: nx.Graph,
    analysis: dict[str, Any],
    output_path: Path,
    now: datetime,
    config: dict[str, Any],
    tz: ZoneInfo,
    alliance_powers: dict[str, AlliancePower],
    phase3_simulation: dict[str, Any] | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    critical_ids = {item["node_id"] for item in analysis["critical_nodes"]}
    warning_hours = float(config.get("protect_warning_hours", 6))
    visual_scale = float(config.get("visual_scale", 1.0))
    font_size = int(config.get("font_size", 18))
    owner_affiliations = build_owner_affiliations(graph, config)

    network = Network(height="100vh", width="100%", bgcolor="#111827", font_color="#f9fafb", cdn_resources="local")
    network.toggle_physics(False)

    for node_id, node_data in graph.nodes(data=True):
        node = Node(**{field: node_data[field] for field in Node.__dataclass_fields__})
        protect = protection_status(node, now, warning_hours, tz)
        border = "#f9fafb"
        if protect["status"] == "expired":
            border = "#ef4444"
        elif protect["status"] == "soon":
            border = "#facc15"
        owner_label = display_node_owner(node)
        power_payload = alliance_power_payload(node.owner, alliance_powers)
        affiliation = strategic_affiliation(node, owner_affiliations)
        power_label = power_payload["power_label"] if power_payload else "-"
        title = f"{node.area} {node.name} / {node.type} / {owner_label} / {power_label}"
        visual_x, visual_y = visual_position(node, config)
        network.add_node(
            node.id,
            label=node.name,
            title=title,
            x=visual_x * visual_scale,
            y=-visual_y * visual_scale,
            physics=False,
            fixed=False,
            size=visual_node_size(node, config),
            shape=visual_node_shape(node, config),
            color={"background": strategic_color(node, owner_affiliations), "border": border},
            borderWidth=6 if node_id in critical_ids else 2,
            area=node.area,
            coord=node.name,
            coordLabel=node.name,
            ownerLabel=owner_label,
            nodeType=node.type,
            owner=node.owner,
            displayOwner=owner_label,
            allianceName=power_payload["alliance_name"] if power_payload else "",
            alliancePower=power_payload["power"] if power_payload else None,
            alliancePowerLabel=power_payload["power_label"] if power_payload else "",
            alliancePowerServer=power_payload["server"] if power_payload else "",
            alliancePowerRank=power_payload["overall_rank"] if power_payload else None,
            alliancePowerServerRank=power_payload["server_rank"] if power_payload else None,
            alliancePowerSide=power_payload["side"] if power_payload else "",
            affiliation=affiliation,
            importance=node.importance,
            rawProtectUntil=node.protect_until or "",
            protectUntil=protect["protect_until"] or "",
            protectStatus=protect["status"],
            hoursRemaining=protect["hours_remaining"],
            acquiredAt=node.acquired_at or "",
            mapStatus=node.status,
            memo=node.memo,
        )
        network.node_map[node.id]["font"] = {
            "size": visual_font_size(node, font_size, config),
            "face": "arial",
            "color": "#f9fafb",
            "strokeWidth": 4,
            "strokeColor": "#111827",
        }

    for source, target, edge_data in graph.edges(data=True):
        edge_id = f"{source}|{target}"
        network.add_edge(source, target, id=edge_id, width=1, title=f"weight: {edge_data.get('weight', 1.0)}")

    network.set_options(
        json.dumps(
            {
                "nodes": {
                    "font": {
                        "size": font_size,
                        "face": "arial",
                        "color": "#f9fafb",
                        "strokeWidth": 4,
                        "strokeColor": "#111827",
                    },
                    "shape": "dot",
                },
                "edges": {
                    "color": {"color": "#94a3b8", "highlight": "#facc15"},
                    "smooth": False,
                },
                "physics": {
                    "enabled": False,
                    "stabilization": False,
                },
                "interaction": {
                    "hover": True,
                    "navigationButtons": True,
                    "keyboard": True,
                },
            }
        )
    )
    copy_local_vis_assets(output_path)
    html = network.generate_html(notebook=False)
    html = localize_vis_resources(strip_remote_bootstrap(html))
    html = add_node_info_panel_v3(html, phase3_simulation)
    html = clean_generated_html(html)
    output_path.write_text(html, encoding="utf-8")


def copy_local_vis_assets(output_path: Path) -> None:
    pyvis_root = Path(inspect.getfile(Network)).parent
    source = pyvis_root / "templates" / "lib"
    target = output_path.parent / "lib"
    if target.exists():
        shutil.rmtree(target)
    target.mkdir(parents=True)
    for asset_dir in ("bindings", "tom-select", "vis-9.1.2"):
        shutil.copytree(source / asset_dir, target / asset_dir)


def strip_remote_bootstrap(html: str) -> str:
    html = re.sub(
        r"\s*<link[^>]+bootstrap@5\.0\.0-beta3/dist/css/bootstrap\.min\.css[^>]+>\s*",
        "",
        html,
        flags=re.DOTALL,
    )
    html = re.sub(
        r"\s*<script[^>]+bootstrap@5\.0\.0-beta3/dist/js/bootstrap\.bundle\.min\.js[^>]*></script>\s*",
        "",
        html,
        flags=re.DOTALL,
    )
    return html


def localize_vis_resources(html: str) -> str:
    html = re.sub(
        r"<link[^>]+vis-network\.min\.css[^>]+>",
        '<link rel="stylesheet" href="lib/vis-9.1.2/vis-network.css" />',
        html,
    )
    html = re.sub(
        r"<script[^>]+vis-network\.min\.js[^>]*></script>",
        '<script src="lib/vis-9.1.2/vis-network.min.js"></script>',
        html,
    )
    return html


def add_node_info_panel(html: str) -> str:
    style = """
<style>
  html,
  body {
    width: 100%;
    height: 100%;
    margin: 0;
    overflow: hidden;
    background: #111827;
  }
  .card {
    width: 100% !important;
    height: 100vh !important;
    border: 0 !important;
    background: #111827 !important;
  }
  .card-body {
    height: 100vh !important;
    padding: 0 !important;
  }
  #mynetwork {
    width: 100vw !important;
    height: 100vh !important;
    border: 0 !important;
  }
  #node-info-panel {
    position: fixed;
    right: 16px;
    top: 16px;
    z-index: 10;
    width: min(360px, calc(100vw - 32px));
    max-height: calc(100vh - 32px);
    overflow: auto;
    background: rgba(15, 23, 42, 0.94);
    color: #f8fafc;
    border: 1px solid rgba(148, 163, 184, 0.55);
    border-radius: 8px;
    box-shadow: 0 18px 45px rgba(0, 0, 0, 0.38);
    font-family: Arial, sans-serif;
    padding: 14px 16px;
  }
  #node-info-panel .node-info-empty {
    color: #cbd5e1;
    font-size: 13px;
    line-height: 1.55;
  }
  #node-info-panel h2 {
    margin: 0 0 10px;
    font-size: 18px;
    line-height: 1.25;
  }
  #node-info-panel dl {
    display: grid;
    grid-template-columns: 96px 1fr;
    gap: 7px 10px;
    margin: 0;
    font-size: 13px;
    line-height: 1.35;
  }
  #node-info-panel dt {
    color: #94a3b8;
  }
  #node-info-panel dd {
    margin: 0;
    word-break: break-word;
  }
  #node-info-panel .node-info-memo {
    grid-column: 1 / -1;
    margin-top: 6px;
    padding-top: 8px;
    border-top: 1px solid rgba(148, 163, 184, 0.35);
    white-space: pre-wrap;
    color: #e2e8f0;
  }
</style>
"""
    panel = """
<div id="node-info-panel">
  <div class="node-info-empty">ノードをクリックすると、管理表由来のエリア・座標・種別・連盟・保護情報をここに表示します。</div>
</div>
"""
    script = """
<script>
  (function () {
    function valueOrDash(value) {
      return value === undefined || value === null || value === "" ? "-" : String(value);
    }
    function escapeHtml(value) {
      return valueOrDash(value)
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#039;");
    }
    function row(label, value) {
      return "<dt>" + escapeHtml(label) + "</dt><dd>" + escapeHtml(value) + "</dd>";
    }
    function renderNodeInfo(nodeId) {
      var panel = document.getElementById("node-info-panel");
      if (!panel || !nodeId || !nodes) return;
      var node = nodes.get(nodeId);
      if (!node) return;
      var memo = valueOrDash(node.memo);
      panel.innerHTML =
        "<h2>" + escapeHtml(valueOrDash(node.area) + " " + valueOrDash(node.coord)) + "</h2>" +
        "<dl>" +
        row("位置キー", node.id) +
        row("種別", node.nodeType) +
        row("連盟", node.owner) +
        row("状態", node.mapStatus) +
        row("取得日時", node.acquiredAt) +
        row("保護切れ", node.rawProtectUntil || node.protectUntil) +
        row("残り時間", node.hoursRemaining === null ? "-" : node.hoursRemaining + " h") +
        row("重要度", node.importance) +
        (memo === "-" ? "" : "<dd class=\\"node-info-memo\\">" + escapeHtml(memo) + "</dd>") +
        "</dl>";
    }
    function resetNodeInfo() {
      var panel = document.getElementById("node-info-panel");
      if (panel) {
        panel.innerHTML = '<div class="node-info-empty">ノードをクリックすると、管理表由来のエリア・座標・種別・連盟・保護情報をここに表示します。</div>';
      }
    }
    if (typeof network !== "undefined") {
      window.renderNodeInfo = renderNodeInfo;
      network.on("selectNode", function (params) {
        renderNodeInfo(params.nodes && params.nodes[0]);
      });
      network.on("deselectNode", resetNodeInfo);
    }
  })();
</script>
"""
    if "</head>" in html:
        html = html.replace("</head>", style + "\n</head>", 1)
    if "<body>" in html:
        html = html.replace("<body>", "<body>\n" + panel, 1)
    return html.replace("</body>", script + "\n</body>", 1)


def add_node_info_panel_v2(html: str) -> str:
    empty_text = "Click a node to show area, coordinate, type, alliance, and protection data from the management table."
    style = """
<style>
  html,
  body {
    width: 100%;
    height: 100%;
    margin: 0;
    overflow: hidden;
    background: #111827;
  }
  .card {
    width: 100% !important;
    height: 100vh !important;
    border: 0 !important;
    background: #111827 !important;
  }
  .card-body {
    height: 100vh !important;
    padding: 0 !important;
  }
  #mynetwork {
    width: 100vw !important;
    height: 100vh !important;
    border: 0 !important;
  }
  #node-info-panel {
    position: fixed;
    right: 16px;
    top: 16px;
    z-index: 10;
    width: min(380px, calc(100vw - 32px));
    max-height: calc(100vh - 32px);
    overflow: auto;
    background: rgba(15, 23, 42, 0.94);
    color: #f8fafc;
    border: 1px solid rgba(148, 163, 184, 0.55);
    border-radius: 8px;
    box-shadow: 0 18px 45px rgba(0, 0, 0, 0.38);
    font-family: Arial, sans-serif;
    padding: 14px 16px;
  }
  #node-info-panel .node-info-empty {
    color: #cbd5e1;
    font-size: 13px;
    line-height: 1.55;
  }
  #node-info-panel h2 {
    margin: 0 0 10px;
    font-size: 18px;
    line-height: 1.25;
  }
  #node-info-panel dl {
    display: grid;
    grid-template-columns: 104px 1fr;
    gap: 7px 10px;
    margin: 0;
    font-size: 13px;
    line-height: 1.35;
  }
  #node-info-panel dt {
    color: #94a3b8;
  }
  #node-info-panel dd {
    margin: 0;
    word-break: break-word;
  }
  #node-info-panel .node-info-memo {
    grid-column: 1 / -1;
    margin-top: 6px;
    padding-top: 8px;
    border-top: 1px solid rgba(148, 163, 184, 0.35);
    white-space: pre-wrap;
    color: #e2e8f0;
  }
  #map-label-toggle {
    position: fixed;
    left: 16px;
    top: 16px;
    z-index: 11;
    min-width: 124px;
    height: 34px;
    padding: 0 12px;
    border: 1px solid rgba(148, 163, 184, 0.65);
    border-radius: 6px;
    background: rgba(15, 23, 42, 0.92);
    color: #f8fafc;
    font: 700 13px Arial, sans-serif;
    cursor: pointer;
  }
</style>
"""
    panel = f"""
<button id="map-label-toggle" type="button">連盟名表示</button>
<div id="node-info-panel">
  <div class="node-info-empty">{empty_text}</div>
</div>
"""
    script = f"""
<script>
  (function () {{
    var emptyText = {json.dumps(empty_text)};
    function valueOrDash(value) {{
      return value === undefined || value === null || value === "" ? "-" : String(value);
    }}
    function escapeHtml(value) {{
      return valueOrDash(value)
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#039;");
    }}
    function row(label, value) {{
      return "<dt>" + escapeHtml(label) + "</dt><dd>" + escapeHtml(value) + "</dd>";
    }}
    function renderNodeInfo(nodeId) {{
      var panel = document.getElementById("node-info-panel");
      if (!panel || !nodeId || !nodes) return;
      var node = nodes.get(nodeId);
      if (!node) return;
      var memo = valueOrDash(node.memo);
      panel.innerHTML =
        "<h2>" + escapeHtml(valueOrDash(node.area) + " " + valueOrDash(node.coord)) + "</h2>" +
        "<dl>" +
        row("Position key", node.id) +
        row("Type", node.nodeType) +
        row("Alliance", node.owner) +
        row("Status", node.mapStatus) +
        row("Acquired", node.acquiredAt) +
        row("Protect until", node.rawProtectUntil || node.protectUntil) +
        row("Hours left", node.hoursRemaining === null ? "-" : node.hoursRemaining + " h") +
        row("Importance", node.importance) +
        (memo === "-" ? "" : "<dd class=\\"node-info-memo\\">" + escapeHtml(memo) + "</dd>") +
        "</dl>";
    }}
    function resetNodeInfo() {{
      var panel = document.getElementById("node-info-panel");
      if (panel) {{
        panel.innerHTML = '<div class="node-info-empty">' + escapeHtml(emptyText) + '</div>';
      }}
    }}
    if (typeof network !== "undefined") {{
      window.renderNodeInfo = renderNodeInfo;
      network.on("selectNode", function (params) {{
        renderNodeInfo(params.nodes && params.nodes[0]);
      }});
      network.on("deselectNode", resetNodeInfo);
    }}
  }})();
</script>
"""
    if "</head>" in html:
        html = html.replace("</head>", style + "\n</head>", 1)
    if "<body>" in html:
        html = html.replace("<body>", "<body>\n" + panel, 1)
    return html.replace("</body>", script + "\n</body>", 1)


def add_node_info_panel_v3(html: str, phase3_simulation: dict[str, Any] | None = None) -> str:
    empty_text = "Click a node to show area, coordinate, type, alliance, and protection data from the management table."
    phase3_json = json.dumps(phase3_simulation or {}, ensure_ascii=False)
    style = """
<style>
  html,
  body {
    width: 100%;
    height: 100%;
    margin: 0;
    overflow: hidden;
    background: #111827;
  }
  .card {
    width: 100% !important;
    height: 100vh !important;
    border: 0 !important;
    background: #111827 !important;
  }
  .card-body {
    height: 100vh !important;
    padding: 0 !important;
  }
  #mynetwork {
    width: 100vw !important;
    height: 100vh !important;
    border: 0 !important;
  }
  #node-info-panel {
    position: fixed;
    right: 16px;
    top: 16px;
    z-index: 10;
    width: min(380px, calc(100vw - 32px));
    max-height: calc(100vh - 32px);
    overflow: auto;
    background: rgba(15, 23, 42, 0.94);
    color: #f8fafc;
    border: 1px solid rgba(148, 163, 184, 0.55);
    border-radius: 8px;
    box-shadow: 0 18px 45px rgba(0, 0, 0, 0.38);
    font-family: Arial, sans-serif;
    padding: 14px 16px;
  }
  #node-info-panel .node-info-empty {
    color: #cbd5e1;
    font-size: 13px;
    line-height: 1.55;
  }
  #node-info-panel h2 {
    margin: 0 0 10px;
    font-size: 18px;
    line-height: 1.25;
  }
  #node-info-panel dl {
    display: grid;
    grid-template-columns: 104px 1fr;
    gap: 7px 10px;
    margin: 0;
    font-size: 13px;
    line-height: 1.35;
  }
  #node-info-panel dt {
    color: #94a3b8;
  }
  #node-info-panel dd {
    margin: 0;
    word-break: break-word;
  }
  #node-info-panel .node-info-memo {
    grid-column: 1 / -1;
    margin-top: 6px;
    padding-top: 8px;
    border-top: 1px solid rgba(148, 163, 184, 0.35);
    white-space: pre-wrap;
    color: #e2e8f0;
  }
  #phase3-score-panel {
    position: fixed;
    left: 16px;
    bottom: 16px;
    z-index: 11;
    width: min(420px, calc(100vw - 32px));
    max-height: min(48vh, 520px);
    overflow: auto;
    padding: 12px;
    border: 1px solid rgba(248, 113, 113, 0.62);
    border-radius: 8px;
    background: rgba(15, 23, 42, 0.94);
    color: #f8fafc;
    box-shadow: 0 18px 45px rgba(0, 0, 0, 0.38);
    font-family: Arial, sans-serif;
  }
  #phase3-score-panel h2 {
    margin: 0 0 8px;
    font-size: 15px;
    line-height: 1.3;
  }
  #phase3-score-panel .phase3-subtitle {
    margin: 0 0 8px;
    color: #cbd5e1;
    font-size: 12px;
    line-height: 1.35;
  }
  #phase3-score-panel .phase3-empty {
    color: #cbd5e1;
    font-size: 13px;
    line-height: 1.45;
  }
  #phase3-score-panel .phase3-item {
    width: 100%;
    margin: 6px 0;
    padding: 8px;
    border: 1px solid rgba(148, 163, 184, 0.45);
    border-radius: 6px;
    background: rgba(30, 41, 59, 0.86);
    color: #f8fafc;
    text-align: left;
    cursor: pointer;
  }
  #phase3-score-panel .phase3-item:hover {
    border-color: rgba(248, 113, 113, 0.86);
    background: rgba(51, 65, 85, 0.92);
  }
  #phase3-score-panel .phase3-main {
    display: flex;
    justify-content: space-between;
    gap: 10px;
    align-items: baseline;
    font-size: 13px;
    font-weight: 700;
  }
  #phase3-score-panel .phase3-score {
    color: #fca5a5;
    white-space: nowrap;
  }
  #phase3-score-panel .phase3-route {
    margin-top: 3px;
    color: #cbd5e1;
    font-size: 12px;
  }
  #phase3-score-panel .phase3-reasons {
    margin-top: 6px;
    display: flex;
    flex-wrap: wrap;
    gap: 4px;
  }
  #phase3-score-panel .phase3-reason {
    padding: 2px 6px;
    border-radius: 999px;
    background: rgba(248, 113, 113, 0.16);
    color: #fecaca;
    font-size: 11px;
    line-height: 1.4;
  }
  #map-toolbar {
    position: fixed;
    left: 16px;
    top: 16px;
    z-index: 12;
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    max-width: calc(100vw - 32px);
  }
  #map-toolbar button {
    min-width: 104px;
    height: 34px;
    padding: 0 12px;
    border: 1px solid rgba(148, 163, 184, 0.65);
    border-radius: 6px;
    background: rgba(15, 23, 42, 0.92);
    color: #f8fafc;
    font: 700 13px Arial, sans-serif;
    cursor: pointer;
  }
  #map-label-toggle {
    min-width: 124px;
  }
  #map-refresh-sheet {
    min-width: 124px;
    border: 1px solid rgba(14, 165, 233, 0.9);
    background: rgba(3, 105, 161, 0.94);
  }
  #map-refresh-sheet:disabled {
    cursor: wait;
    opacity: 0.65;
  }
  #map-route-mode {
    border: 1px solid rgba(168, 85, 247, 0.9);
    background: rgba(88, 28, 135, 0.94);
  }
  #map-legend {
    position: fixed;
    left: 16px;
    top: 62px;
    z-index: 10;
    width: 250px;
    padding: 10px 12px;
    border: 1px solid rgba(148, 163, 184, 0.55);
    border-radius: 8px;
    background: rgba(15, 23, 42, 0.9);
    color: #e5e7eb;
    font: 12px/1.35 Arial, sans-serif;
  }
  #map-legend h3 {
    margin: 0 0 8px;
    color: #f8fafc;
    font-size: 13px;
  }
  #map-legend .legend-row {
    display: flex;
    align-items: center;
    gap: 8px;
    margin: 5px 0;
  }
  #map-legend .legend-dot {
    width: 14px;
    height: 14px;
    border-radius: 50%;
    border: 2px solid #111827;
    flex: 0 0 auto;
  }
  #map-legend .legend-border {
    width: 18px;
    height: 10px;
    border-radius: 8px;
    background: transparent;
    flex: 0 0 auto;
  }
  @media (max-width: 900px) {
    #map-toolbar {
      max-width: calc(100vw - 32px);
    }
    #map-legend {
      top: 150px;
    }
  }
</style>
"""
    panel = f"""
<div id="map-toolbar">
<button id="map-label-toggle" type="button">&#36899;&#30431;&#21517;&#34920;&#31034;</button>
<button id="map-reset-layout" type="button">&#20301;&#32622;&#12522;&#12475;&#12483;&#12488;</button>
<button id="map-highlight-boundary" type="button">&#22659;&#30028;&#24375;&#35519;</button>
<button id="map-highlight-boundary-depth" type="button">&#22659;&#30028;+&#20869;&#20596;</button>
<button id="map-highlight-boundary-depth-2" type="button">&#22659;&#30028;+&#20869;&#20596;+&#20869;&#20596;</button>
<button id="map-highlight-boundary-depth-3" type="button">&#22659;&#30028;+&#20869;&#20596;+&#20869;&#20596;+&#20869;&#20596;</button>
<button id="map-highlight-enemy-threat" type="button">&#25973;&#20405;&#25915;&#20104;&#28204;</button>
<button id="map-highlight-pact-threat" type="button">&#21332;&#23450;&#36796;&#12415;&#25973;&#20405;&#25915;&#20104;&#28204;</button>
<button id="map-highlight-friendly-pressure" type="button">&#21619;&#26041;&#20405;&#25915;&#20505;&#35036;</button>
<button id="map-highlight-interdiction" type="button">&#36974;&#26029;&#20505;&#35036;</button>
<button id="map-highlight-risk-avoidance" type="button">&#21361;&#38522;&#22238;&#36991;</button>
<button id="map-phase3-defense" type="button">Phase3防衛TOP</button>
<button id="map-phase3-attack" type="button">Phase3攻撃TOP</button>
<button id="map-phase3-interdiction" type="button">Phase3遮断TOP</button>
<button id="map-phase3-risk" type="button">Phase3危険TOP</button>
<button id="map-phase3-protection" type="button">Phase3保護TOP</button>
<button id="map-phase3-time" type="button">Phase3時間TOP</button>
<button id="map-phase3-overall" type="button">Phase3総合TOP</button>
<button id="map-highlight-enemy-expand" type="button">&#25973;&#26410;&#21462;&#24471;&#25313;&#24373;</button>
<button id="map-highlight-friendly-expand" type="button">&#21619;&#26041;&#26410;&#21462;&#24471;&#25313;&#24373;</button>
<button id="map-clear-edge-highlight" type="button">&#24375;&#35519;&#35299;&#38500;</button>
<button id="map-refresh-sheet" type="button">&#12510;&#12483;&#12503;&#26368;&#26032;&#21270;</button>
<button id="map-route-ignore-power" type="button">&#25126;&#21147;&#28961;&#35222;&#12523;&#12540;&#12488;</button>
<button id="map-route-weaker-power" type="button">&#22987;&#28857;&#12424;&#12426;&#20302;&#25126;&#21147;&#36890;&#36942;</button>
<button id="map-clear-route" type="button">&#12523;&#12540;&#12488;&#35299;&#38500;</button>
</div>
<div id="map-legend" aria-label="legend">
  <h3>凡例</h3>
  <div class="legend-row"><span class="legend-dot" style="background:#2563eb"></span><span>#534 / JDX</span></div>
  <div class="legend-row"><span class="legend-dot" style="background:#16a34a"></span><span>味方</span></div>
  <div class="legend-row"><span class="legend-dot" style="background:#dc2626"></span><span>敵</span></div>
  <div class="legend-row"><span class="legend-dot" style="background:#f8fafc"></span><span>未取得</span></div>
  <div class="legend-row"><span class="legend-dot" style="background:#6b7280"></span><span>破壊</span></div>
  <div class="legend-row"><span class="legend-border" style="border:3px solid #ef4444"></span><span>保護終了済み</span></div>
  <div class="legend-row"><span class="legend-border" style="border:3px solid #facc15"></span><span>保護終了が近い</span></div>
</div>
<div id="node-info-panel">
  <div class="node-info-empty">{empty_text}</div>
</div>
<div id="phase3-score-panel">
  <h2>Phase3スコア</h2>
  <div class="phase3-empty">攻撃TOP・遮断TOP・危険TOPを押すと、state.jsonのスコア根拠を表示します。</div>
</div>
"""
    attach_js = f"""
                  (function () {{
                    var emptyText = {json.dumps(empty_text)};
                    var phase3Simulation = {phase3_json};
                    function valueOrDash(value) {{
                      return value === undefined || value === null || value === "" ? "-" : String(value);
                    }}
                    function escapeHtml(value) {{
                      return valueOrDash(value)
                        .replaceAll("&", "&amp;")
                        .replaceAll("<", "&lt;")
                        .replaceAll(">", "&gt;")
                        .replaceAll('"', "&quot;")
                        .replaceAll("'", "&#039;");
                    }}
                    function row(label, value) {{
                      return "<dt>" + escapeHtml(label) + "</dt><dd>" + escapeHtml(value) + "</dd>";
                    }}
                    var initialPositions = {{}};
                    if (nodes) {{
                      nodes.get().forEach(function (node) {{
                        initialPositions[node.id] = {{ x: node.x, y: node.y }};
                      }});
                    }}
                    var labelMode = "coord";
                    function applyLabelMode() {{
                      if (!nodes) return;
                      var updates = nodes.get().map(function (node) {{
                        return {{
                          id: node.id,
                          label: labelMode === "owner"
                            ? (
                                node.nodeType === "\\u4ea4\\u6613\\u5730"
                                  ? "\\u4ea4\\u6613\\u5730"
                                  : valueOrDash(node.ownerLabel || node.displayOwner)
                              )
                            : valueOrDash(node.coordLabel || node.coord)
                        }};
                      }});
                      nodes.update(updates);
                      var button = document.getElementById("map-label-toggle");
                      if (button) {{
                        button.textContent = labelMode === "owner" ? "\\u5ea7\\u6a19\\u8868\\u793a" : "\\u9023\\u76df\\u540d\\u8868\\u793a";
                      }}
                    }}
                    var labelButton = document.getElementById("map-label-toggle");
                    if (labelButton) {{
                      labelButton.addEventListener("click", function () {{
                        labelMode = labelMode === "owner" ? "coord" : "owner";
                        applyLabelMode();
                      }});
                    }}
                    var resetButton = document.getElementById("map-reset-layout");
                    if (resetButton) {{
                      resetButton.addEventListener("click", function () {{
                        if (!nodes || !network) return;
                        var updates = Object.keys(initialPositions).map(function (id) {{
                          return {{ id: id, x: initialPositions[id].x, y: initialPositions[id].y }};
                        }});
                        nodes.update(updates);
                        network.fit({{ animation: {{ duration: 180, easingFunction: "easeInOutQuad" }} }});
                      }});
                    }}
                    var defaultEdgeStyle = {{ color: {{ color: "#94a3b8", highlight: "#facc15" }}, width: 1 }};
                    var routeMode = null;
                    var routeStartNodeId = null;
                    var routeEdgeIds = [];
                    function updateRouteButtonText() {{
                      var ignoreButton = document.getElementById("map-route-ignore-power");
                      var weakerButton = document.getElementById("map-route-weaker-power");
                      var buttons = [
                        {{ element: ignoreButton, mode: "ignore", label: "\\u6226\\u529b\\u7121\\u8996\\u30eb\\u30fc\\u30c8" }},
                        {{ element: weakerButton, mode: "weaker", label: "\\u59cb\\u70b9\\u3088\\u308a\\u4f4e\\u6226\\u529b\\u901a\\u904e" }}
                      ];
                      buttons.forEach(function (button) {{
                        if (!button.element) return;
                        if (routeMode !== button.mode) {{
                          button.element.textContent = button.label;
                          button.element.style.outline = "";
                        }} else if (routeStartNodeId) {{
                          button.element.textContent = "\\u7d42\\u70b9\\u9078\\u629e\\u4e2d";
                          button.element.style.outline = "2px solid #facc15";
                        }} else {{
                          button.element.textContent = "\\u59cb\\u70b9\\u9078\\u629e\\u4e2d";
                          button.element.style.outline = "2px solid #facc15";
                        }}
                      }});
                    }}
                    function startRouteMode(mode) {{
                      routeMode = mode;
                      routeStartNodeId = null;
                      routeEdgeIds = [];
                      resetEdgeHighlights();
                      if (network) network.unselectAll();
                      updateRouteButtonText();
                    }}
                    function routeStartAlliancePowerValue() {{
                      if (!nodes || !routeStartNodeId) return 0;
                      var startNode = nodes.get(routeStartNodeId);
                      return Number(startNode && startNode.alliancePower ? startNode.alliancePower : 0);
                    }}
                    function canUseRouteNode(node, mode, isEndpoint) {{
                      if (!isFisheryNode(node)) return false;
                      if (mode !== "weaker" || isEndpoint) return true;
                      if (node.affiliation !== "enemy") return true;
                      var selfPower = routeStartAlliancePowerValue();
                      var enemyPower = Number(node.alliancePower || 0);
                      if (!selfPower || !enemyPower) return false;
                      return enemyPower < selfPower;
                    }}
                    function routeColor() {{
                      return routeMode === "weaker" ? "#10b981" : "#a855f7";
                    }}
                    function routeNoPathMessage() {{
                      if (routeMode === "weaker") {{
                        return "\\u9078\\u629e\\u3057\\u305f2\\u30ce\\u30fc\\u30c9\\u9593\\u306b\\u3001\\u59cb\\u70b9\\u9023\\u76df\\u3088\\u308a\\u4f4e\\u6226\\u529b\\u306e\\u6575\\u9023\\u76df\\u3060\\u3051\\u3092\\u901a\\u308b\\u30eb\\u30fc\\u30c8\\u304c\\u3042\\u308a\\u307e\\u305b\\u3093\\u3002";
                      }}
                      return "\\u9078\\u629e\\u3057\\u305f2\\u30ce\\u30fc\\u30c9\\u9593\\u306e\\u30eb\\u30fc\\u30c8\\u304c\\u3042\\u308a\\u307e\\u305b\\u3093\\u3002";
                    }}
                    function clearRouteState() {{
                      routeMode = null;
                      routeStartNodeId = null;
                      routeEdgeIds = [];
                      updateRouteButtonText();
                    }}
                    function resetEdgeHighlights() {{
                      if (!edges) return;
                      edges.update(edges.get().map(function (edge) {{
                        return {{
                          id: edge.id,
                          color: defaultEdgeStyle.color,
                          width: defaultEdgeStyle.width
                        }};
                      }}));
                    }}
                    function clearAllEdgeHighlights() {{
                      clearRouteState();
                      resetEdgeHighlights();
                    }}
                    function highlightEdges(edgeIds, color, width) {{
                      if (!edges || !edgeIds.length) return;
                      edges.update(edgeIds.map(function (edgeId) {{
                        return {{
                          id: edgeId,
                          color: {{ color: color, highlight: color }},
                          width: width
                        }};
                      }}));
                    }}
                    function isFisheryNode(node) {{
                      return Boolean(node && node.nodeType === "\\u6f01\\u5834");
                    }}
                    function isCityNode(node) {{
                      return Boolean(node && node.nodeType === "\\u90fd\\u5e02");
                    }}
                    function isFriendlyAffiliation(affiliation) {{
                      return affiliation === "self" || affiliation === "ally";
                    }}
                    function isFriendlyInteriorAffiliation(affiliation) {{
                      return isFriendlyAffiliation(affiliation) || affiliation === "unowned";
                    }}
                    function adjacentFisheryIds(nodeId) {{
                      if (!edges || !nodes || !nodeId) return [];
                      var found = {{}};
                      edges.get().forEach(function (edge) {{
                        var otherId = null;
                        if (edge.from === nodeId) {{
                          otherId = edge.to;
                        }} else if (edge.to === nodeId) {{
                          otherId = edge.from;
                        }}
                        if (!otherId) return;
                        var other = nodes.get(otherId);
                        if (isFisheryNode(other)) found[otherId] = true;
                      }});
                      return Object.keys(found);
                    }}
                    function routeEndpointCandidates(nodeId) {{
                      var node = nodes.get(nodeId);
                      if (isFisheryNode(node)) return [nodeId];
                      if (isCityNode(node)) return adjacentFisheryIds(nodeId);
                      return [];
                    }}
                    function shortestRouteEdges(startId, targetId, mode) {{
                      if (!edges || !nodes || !startId || !targetId) return null;
                      var startCandidates = routeEndpointCandidates(startId);
                      var targetCandidates = routeEndpointCandidates(targetId);
                      if (!startCandidates.length || !targetCandidates.length) return null;
                      var adjacency = {{}};
                      var targetSet = {{}};
                      targetCandidates.forEach(function (nodeId) {{ targetSet[nodeId] = true; }});
                      var byPair = {{}};
                      edges.get().forEach(function (edge) {{
                        var source = nodes.get(edge.from);
                        var target = nodes.get(edge.to);
                        if (!isFisheryNode(source) || !isFisheryNode(target)) return;
                        if (!canUseRouteNode(source, mode, Boolean(targetSet[edge.from])) && startCandidates.indexOf(edge.from) === -1) return;
                        if (!canUseRouteNode(target, mode, Boolean(targetSet[edge.to])) && startCandidates.indexOf(edge.to) === -1) return;
                        if (!adjacency[edge.from]) adjacency[edge.from] = [];
                        if (!adjacency[edge.to]) adjacency[edge.to] = [];
                        adjacency[edge.from].push(edge.to);
                        adjacency[edge.to].push(edge.from);
                        byPair[edge.from + "\\u0000" + edge.to] = edge.id;
                        byPair[edge.to + "\\u0000" + edge.from] = edge.id;
                      }});
                      var queue = [];
                      var distance = {{}};
                      var predecessors = {{}};
                      var routeCounts = {{}};
                      var targetDistance = null;
                      startCandidates.forEach(function (nodeId) {{
                        if (distance[nodeId] !== undefined) return;
                        distance[nodeId] = 0;
                        routeCounts[nodeId] = 1;
                        queue.push(nodeId);
                        if (targetSet[nodeId]) targetDistance = 0;
                      }});
                      for (var i = 0; i < queue.length; i += 1) {{
                        var current = queue[i];
                        if (targetDistance !== null && distance[current] >= targetDistance) continue;
                        var neighbors = adjacency[current] || [];
                        for (var j = 0; j < neighbors.length; j += 1) {{
                          var next = neighbors[j];
                          if (!targetSet[next] && !canUseRouteNode(nodes.get(next), mode, false)) continue;
                          var nextDistance = distance[current] + 1;
                          if (distance[next] === undefined) {{
                            distance[next] = nextDistance;
                            predecessors[next] = [current];
                            routeCounts[next] = routeCounts[current];
                            if (targetSet[next]) {{
                              targetDistance = nextDistance;
                            }}
                            queue.push(next);
                          }} else if (distance[next] === nextDistance) {{
                            predecessors[next].push(current);
                            routeCounts[next] = Math.min(9999, (routeCounts[next] || 0) + (routeCounts[current] || 0));
                          }}
                        }}
                      }}
                      if (targetDistance === null) return null;
                      var targetNodes = targetCandidates.filter(function (nodeId) {{
                        return distance[nodeId] === targetDistance;
                      }});
                      var edgeIdSet = {{}};
                      var stack = targetNodes.slice();
                      var seenNodes = {{}};
                      while (stack.length) {{
                        var nodeId = stack.pop();
                        var preds = predecessors[nodeId] || [];
                        for (var k = 0; k < preds.length; k += 1) {{
                          var pred = preds[k];
                          var edgeId = byPair[pred + "\\u0000" + nodeId];
                          if (edgeId) edgeIdSet[edgeId] = true;
                          if (distance[pred] !== 0 && !seenNodes[pred]) {{
                            seenNodes[pred] = true;
                            stack.push(pred);
                          }}
                        }}
                      }}
                      return {{
                        edgeIds: Object.keys(edgeIdSet),
                        distance: targetDistance,
                        routeCount: targetNodes.reduce(function (total, nodeId) {{
                          return Math.min(9999, total + (routeCounts[nodeId] || 0));
                        }}, 0) || 1
                      }};
                    }}
                    function showShortestRoute(startId, targetId) {{
                      var route = shortestRouteEdges(startId, targetId, routeMode || "ignore");
                      resetEdgeHighlights();
                      if (!route) {{
                        alert(routeNoPathMessage());
                        clearRouteState();
                        return;
                      }}
                      if (!route.edgeIds.length) {{
                        alert("\\u59cb\\u70b9\\u304c\\u76ee\\u7684\\u5730\\u90fd\\u5e02\\u306b\\u96a3\\u63a5\\u3059\\u308b\\u6700\\u77ed\\u306e\\u6f01\\u5834\\u3067\\u3059\\u3002");
                        clearRouteState();
                        return;
                      }}
                      routeEdgeIds = route.edgeIds;
                      highlightEdges(routeEdgeIds, routeColor(), 6);
                      if (network) network.unselectAll();
                      clearRouteState();
                    }}
                    function handleRouteSelection(nodeId) {{
                      if (!routeMode || !nodeId) return false;
                      if (!routeStartNodeId) {{
                        routeStartNodeId = nodeId;
                        resetEdgeHighlights();
                        if (network) network.unselectAll();
                        updateRouteButtonText();
                        return true;
                      }}
                      showShortestRoute(routeStartNodeId, nodeId);
                      return true;
                    }}
                    function highlightConnectedEdges(nodeId) {{
                      if (!nodes || !edges || !nodeId) return;
                      var node = nodes.get(nodeId);
                      if (!node || node.nodeType !== "\\u6f01\\u5834") return;
                      clearRouteState();
                      resetEdgeHighlights();
                      var selectedEdges = edges.get().filter(function (edge) {{
                        return edge.from === nodeId || edge.to === nodeId;
                      }}).map(function (edge) {{ return edge.id; }});
                      highlightEdges(selectedEdges, "#38bdf8", 4);
                    }}
                    function highlightSelfEnemyFisheryEdges() {{
                      if (!nodes || !edges) return;
                      clearRouteState();
                      resetEdgeHighlights();
                      var boundary = collectBoundaryFisheryEdges();
                      highlightEdges(boundary.unownedBoundaryEdgeIds, "#facc15", 4);
                      highlightEdges(boundary.friendlyBoundaryEdgeIds, "#fb923c", 5);
                    }}
                    function collectBoundaryFisheryEdges() {{
                      var friendlyBoundaryEdgeIds = [];
                      var unownedBoundaryEdgeIds = [];
                      var boundaryFriendlyNodes = {{}};
                      edges.get().forEach(function (edge) {{
                        var source = nodes.get(edge.from);
                        var target = nodes.get(edge.to);
                        if (!source || !target) return;
                        if (source.nodeType !== "\\u6f01\\u5834" || target.nodeType !== "\\u6f01\\u5834") return;
                        if (
                          (isFriendlyAffiliation(source.affiliation) && target.affiliation === "enemy") ||
                          (source.affiliation === "enemy" && isFriendlyAffiliation(target.affiliation))
                        ) {{
                          friendlyBoundaryEdgeIds.push(edge.id);
                          if (isFriendlyAffiliation(source.affiliation)) {{
                            boundaryFriendlyNodes[source.id] = true;
                          }} else {{
                            boundaryFriendlyNodes[target.id] = true;
                          }}
                        }} else if (
                          (source.affiliation === "enemy" && target.affiliation === "unowned") ||
                          (source.affiliation === "unowned" && target.affiliation === "enemy")
                        ) {{
                          unownedBoundaryEdgeIds.push(edge.id);
                        }}
                      }});
                      return {{
                        friendlyBoundaryEdgeIds: friendlyBoundaryEdgeIds,
                        unownedBoundaryEdgeIds: unownedBoundaryEdgeIds,
                        boundaryFriendlyNodes: boundaryFriendlyNodes
                      }};
                    }}
                    function collectInteriorDepthEdgeIds(boundaryFriendlyNodes, maxDepth) {{
                      var interiorEdges = [];
                      var adjacency = {{}};
                      edges.get().forEach(function (edge) {{
                        var source = nodes.get(edge.from);
                        var target = nodes.get(edge.to);
                        if (!source || !target) return;
                        if (source.nodeType !== "\\u6f01\\u5834" || target.nodeType !== "\\u6f01\\u5834") return;
                        if (!isFriendlyInteriorAffiliation(source.affiliation) || !isFriendlyInteriorAffiliation(target.affiliation)) return;
                        interiorEdges.push(edge);
                        if (!adjacency[edge.from]) adjacency[edge.from] = [];
                        if (!adjacency[edge.to]) adjacency[edge.to] = [];
                        adjacency[edge.from].push(edge.to);
                        adjacency[edge.to].push(edge.from);
                      }});
                      var distances = {{}};
                      var queue = [];
                      Object.keys(boundaryFriendlyNodes).forEach(function (nodeId) {{
                        distances[nodeId] = 0;
                        queue.push(nodeId);
                      }});
                      for (var i = 0; i < queue.length; i += 1) {{
                        var nodeId = queue[i];
                        var distance = distances[nodeId];
                        if (distance >= maxDepth) continue;
                        var neighbors = adjacency[nodeId] || [];
                        for (var j = 0; j < neighbors.length; j += 1) {{
                          var next = neighbors[j];
                          if (distances[next] === undefined) {{
                            distances[next] = distance + 1;
                            queue.push(next);
                          }}
                        }}
                      }}
                      return interiorEdges.filter(function (edge) {{
                        var sourceDistance = distances[edge.from];
                        var targetDistance = distances[edge.to];
                        if (sourceDistance === undefined || targetDistance === undefined) return false;
                        return Math.min(sourceDistance, targetDistance) < maxDepth && Math.max(sourceDistance, targetDistance) <= maxDepth;
                      }}).map(function (edge) {{ return edge.id; }});
                    }}
                    function highlightSelfEnemyFisheryEdgesWithSelfDepth(maxDepth) {{
                      if (!nodes || !edges) return;
                      clearRouteState();
                      resetEdgeHighlights();
                      var boundary = collectBoundaryFisheryEdges();
                      var friendlyDepthEdgeIds = collectInteriorDepthEdgeIds(boundary.boundaryFriendlyNodes, maxDepth || 1);
                      highlightEdges(friendlyDepthEdgeIds, "#f43f5e", 4);
                      highlightEdges(boundary.unownedBoundaryEdgeIds, "#facc15", 4);
                      highlightEdges(boundary.friendlyBoundaryEdgeIds, "#fb923c", 5);
                    }}
                    function scoreCandidateEdge(source, target) {{
                      var sourcePower = Number(source.alliancePower || 0);
                      var targetPower = Number(target.alliancePower || 0);
                      var powerScore = sourcePower && targetPower ? sourcePower / targetPower : sourcePower / 1000000000;
                      return powerScore + Number(target.importance || 0) / 10;
                    }}
                    function adjacentNodes(nodeId, predicate) {{
                      var results = [];
                      if (!edges || !nodes) return results;
                      edges.get().forEach(function (edge) {{
                        var otherId = null;
                        if (edge.from === nodeId) {{
                          otherId = edge.to;
                        }} else if (edge.to === nodeId) {{
                          otherId = edge.from;
                        }}
                        if (!otherId) return;
                        var other = nodes.get(otherId);
                        if (other && predicate(other)) results.push(other);
                      }});
                      return results;
                    }}
                    function enemyFisheryDensity(node) {{
                      return adjacentNodes(node.id, function (other) {{
                        return isFisheryNode(other) && other.affiliation === "enemy";
                      }}).length;
                    }}
                    function adjacentEnemyCityValue(node) {{
                      return adjacentNodes(node.id, function (other) {{
                        return isCityNode(other) && other.affiliation === "enemy";
                      }}).reduce(function (total, city) {{
                        return total + Number(city.importance || 0);
                      }}, 0);
                    }}
                    function scoreAttackCandidate(attacker, defender) {{
                      var sourcePower = Number(attacker.alliancePower || 0);
                      var targetPower = Number(defender.alliancePower || 0);
                      var powerRatio = sourcePower && targetPower ? sourcePower / targetPower : sourcePower / 1000000000;
                      var enemyPowerRatio = sourcePower && targetPower ? targetPower / sourcePower : 0;
                      return (
                        Number(defender.importance || 0) * 1.2 +
                        adjacentEnemyCityValue(defender) * 0.9 +
                        powerRatio * 2.0 -
                        enemyFisheryDensity(defender) * 0.35 -
                        Math.max(0, enemyPowerRatio - 1) * 2.5
                      );
                    }}
                    function scoreInterdictionCandidate(attacker, city) {{
                      var targetPower = Number(city.alliancePower || 0);
                      var targetDegree = adjacentNodes(city.id, function () {{ return true; }}).length;
                      var friendlyNeighbors = adjacentNodes(city.id, function (other) {{
                        return isFisheryNode(other) && isFriendlyAffiliation(other.affiliation);
                      }}).length;
                      var enemyNeighbors = adjacentNodes(city.id, function (other) {{
                        return isFisheryNode(other) && other.affiliation === "enemy";
                      }}).length;
                      return (
                        Number(city.importance || 0) * 2.0 +
                        targetDegree * 0.8 +
                        enemyNeighbors * 0.8 +
                        friendlyNeighbors * 0.6 +
                        targetPower / 10000000000
                      );
                    }}
                    function scoreRiskCandidate(friendly, enemy) {{
                      var friendlyPower = Number(friendly.alliancePower || 0);
                      var enemyPower = Number(enemy.alliancePower || 0);
                      var enemyRatio = friendlyPower && enemyPower ? enemyPower / friendlyPower : enemyPower / 1000000000;
                      return (
                        enemyRatio * 3.0 +
                        enemyFisheryDensity(enemy) * 0.8 +
                        adjacentEnemyCityValue(enemy) * 0.15 +
                        Number(enemy.importance || 0) * 0.5
                      );
                    }}
                    function candidateEdgeRecords(kind, maxItems) {{
                      var records = [];
                      edges.get().forEach(function (edge) {{
                        var source = nodes.get(edge.from);
                        var target = nodes.get(edge.to);
                        if (!source || !target) return;
                        function pushRecord(attacker, defender, score) {{
                          records.push({{
                            id: edge.id,
                            score: score === undefined ? scoreCandidateEdge(attacker, defender) : score,
                            attacker: attacker.id,
                            defender: defender.id
                          }});
                        }}
                        if (kind === "interdiction") {{
                          if (isFisheryNode(source) && isFriendlyAffiliation(source.affiliation) && isCityNode(target) && target.affiliation === "enemy") {{
                            pushRecord(source, target, scoreInterdictionCandidate(source, target));
                          }}
                          if (isFisheryNode(target) && isFriendlyAffiliation(target.affiliation) && isCityNode(source) && source.affiliation === "enemy") {{
                            pushRecord(target, source, scoreInterdictionCandidate(target, source));
                          }}
                          return;
                        }}
                        if (!isFisheryNode(source) || !isFisheryNode(target)) return;
                        if (kind === "enemyThreat") {{
                          if (source.affiliation === "enemy" && isFriendlyAffiliation(target.affiliation)) pushRecord(source, target, scoreRiskCandidate(target, source));
                          if (target.affiliation === "enemy" && isFriendlyAffiliation(source.affiliation)) pushRecord(target, source, scoreRiskCandidate(source, target));
                        }} else if (kind === "friendlyPressure") {{
                          if (isFriendlyAffiliation(source.affiliation) && target.affiliation === "enemy") pushRecord(source, target, scoreAttackCandidate(source, target));
                          if (isFriendlyAffiliation(target.affiliation) && source.affiliation === "enemy") pushRecord(target, source, scoreAttackCandidate(target, source));
                        }} else if (kind === "riskAvoidance") {{
                          if (isFriendlyAffiliation(source.affiliation) && target.affiliation === "enemy") pushRecord(source, target, scoreRiskCandidate(source, target));
                          if (isFriendlyAffiliation(target.affiliation) && source.affiliation === "enemy") pushRecord(target, source, scoreRiskCandidate(target, source));
                        }} else if (kind === "enemyExpand") {{
                          if (source.affiliation === "enemy" && target.affiliation === "unowned") pushRecord(source, target);
                          if (target.affiliation === "enemy" && source.affiliation === "unowned") pushRecord(target, source);
                        }} else if (kind === "friendlyExpand") {{
                          if (isFriendlyAffiliation(source.affiliation) && target.affiliation === "unowned") pushRecord(source, target);
                          if (isFriendlyAffiliation(target.affiliation) && source.affiliation === "unowned") pushRecord(target, source);
                        }}
                      }});
                      records.sort(function (left, right) {{
                        if (right.score !== left.score) return right.score - left.score;
                        return left.id < right.id ? -1 : 1;
                      }});
                      return records.slice(0, maxItems || 30);
                    }}
                    function highlightCandidateEdges(kind, color, width) {{
                      if (!nodes || !edges) return;
                      clearRouteState();
                      resetEdgeHighlights();
                      var records = candidateEdgeRecords(kind, 30);
                      highlightEdges(records.map(function (record) {{ return record.id; }}), color, width || 6);
                    }}
                    function phase3CandidateConfig(kind) {{
                    var configs = {{
                        defense: {{
                          title: "Phase3 防衛候補 TOP",
                          subtitle: "分断リスク・味方防衛線・中央接続・敵隣接を含む防衛優先度",
                          key: "defense_priorities",
                          color: "#2563eb",
                          width: 8
                        }},
                        attack: {{
                          title: "Phase3 攻撃候補 TOP",
                          subtitle: "敵境界・中央接続・保護終了・反撃リスクを含む攻撃スコア",
                          key: "attack_priorities",
                          color: "#f97316",
                          width: 8
                        }},
                        interdiction: {{
                          title: "Phase3 遮断候補 TOP",
                          subtitle: "都市破壊による到達不能化や隣接遮断を重視",
                          key: "interdiction_priorities",
                          color: "#0ea5e9",
                          width: 8
                        }},
                        risk: {{
                          title: "Phase3 危険回避 TOP",
                          subtitle: "反撃リスク・高戦力敵隣接・敵密集を重視",
                          key: "risk_watchlist",
                          color: "#be123c",
                          width: 8
                        }},
                        pactThreat: {{
                          title: "協定込み敵侵攻予測 TOP",
                          subtitle: "連盟協定で敵が利用し得る一段先の隣接を重視",
                          key: "pact_threat_options",
                          color: "#e11d48",
                          width: 8
                        }},
                        protection: {{
                          title: "Phase3 保護終了 TOP",
                          subtitle: "保護切れ・保護終了間近・保護更新可能性を重視",
                          key: "protection_watchlist",
                          color: "#facc15",
                          width: 8
                        }},
                        time: {{
                          title: "Phase3 時間評価 TOP",
                          subtitle: "水曜/土曜戦闘日・応戦期間・現在時刻との距離を重視",
                          key: "time_sensitive_nodes",
                          color: "#14b8a6",
                          width: 8
                        }},
                        overall: {{
                          title: "Phase3 総合スコア TOP",
                          subtitle: "防衛・攻撃・遮断・危険・保護・時間を統合した上位候補",
                          key: "node_evaluations",
                          color: "#f43f5e",
                          width: 8
                        }}
                      }};
                      return configs[kind] || configs.attack;
                    }}
                    function phase3TargetLabel(record) {{
                      var target = record && record.target ? record.target : {{}};
                      return valueOrDash(target.area || "") + " " + valueOrDash(record.node || target.name || record.node_id);
                    }}
                    function phase3SourceLabel(record) {{
                      var source = record && record.source ? record.source : {{}};
                      return valueOrDash(source.area || "") + " " + valueOrDash(source.name || source.id);
                    }}
                    function phase3RouteText(record) {{
                      if (record && record.source) {{
                        return phase3SourceLabel(record) + ' → ' + phase3TargetLabel(record);
                      }}
                      var labels = [];
                      if (record && record.classification) labels.push("分類: " + record.classification);
                      if (record && record.defense_score !== undefined) labels.push("防衛 " + Number(record.defense_score || 0).toFixed(1));
                      if (record && record.attack_score !== undefined) labels.push("攻撃 " + Number(record.attack_score || 0).toFixed(1));
                      if (record && record.interdiction_score !== undefined) labels.push("遮断 " + Number(record.interdiction_score || 0).toFixed(1));
                      return labels.length ? labels.join(" / ") : phase3TargetLabel(record);
                    }}
                    function phase3ReasonBadges(record) {{
                      var reasons = Array.isArray(record.reasons) ? record.reasons.slice(0, 5) : [];
                      if (!reasons.length) return "";
                      return '<div class="phase3-reasons">' + reasons.map(function (reason) {{
                        return '<span class="phase3-reason">' + escapeHtml(reason) + '</span>';
                      }}).join("") + '</div>';
                    }}
                    function phase3EdgeId(record) {{
                      if (!record) return null;
                      if (record.edge_id && edges && edges.get(record.edge_id)) return record.edge_id;
                      if (Array.isArray(record.edge) && record.edge.length === 2) {{
                        var forward = record.edge[0] + "|" + record.edge[1];
                        var reverse = record.edge[1] + "|" + record.edge[0];
                        if (edges && edges.get(forward)) return forward;
                        if (edges && edges.get(reverse)) return reverse;
                        return record.edge_id || forward;
                      }}
                      return record.edge_id || null;
                    }}
                    function focusPhase3Candidate(record, config) {{
                      if (!record || !nodes || !edges) return;
                      clearRouteState();
                      resetEdgeHighlights();
                      var edgeId = phase3EdgeId(record);
                      if (edgeId) {{
                        highlightEdges([edgeId], config.color, config.width || 8);
                      }}
                      var nodeId = record.node_id || (record.target && record.target.id);
                      if (nodeId && nodes.get(nodeId)) {{
                        network.selectNodes([nodeId]);
                        network.focus(nodeId, {{
                          scale: 1.0,
                          animation: {{ duration: 220, easingFunction: "easeInOutQuad" }}
                        }});
                        renderNodeInfo(nodeId);
                      }}
                    }}
                    function renderPhase3Candidates(kind) {{
                      var panel = document.getElementById("phase3-score-panel");
                      if (!panel) return;
                      var config = phase3CandidateConfig(kind);
                      var records = phase3Simulation && Array.isArray(phase3Simulation[config.key])
                        ? phase3Simulation[config.key].slice(0, 10)
                        : [];
                      if (!records.length) {{
                        panel.innerHTML =
                          '<h2>' + escapeHtml(config.title) + '</h2>' +
                          '<div class="phase3-empty">候補がありません。state.jsonの invasion_simulation を確認してください。</div>';
                        return;
                      }}
                      clearRouteState();
                      resetEdgeHighlights();
                      highlightEdges(records.map(phase3EdgeId).filter(Boolean), config.color, config.width || 8);
                      panel.innerHTML =
                        '<h2>' + escapeHtml(config.title) + '</h2>' +
                        '<p class="phase3-subtitle">' + escapeHtml(config.subtitle) + '</p>' +
                        records.map(function (record, index) {{
                          return '<button class="phase3-item" type="button" data-phase3-index="' + index + '">' +
                            '<div class="phase3-main"><span>' + (index + 1) + '. ' + escapeHtml(phase3TargetLabel(record)) + '</span>' +
                            '<span class="phase3-score">' + escapeHtml(Number(record.score || 0).toFixed(1)) + '</span></div>' +
                            '<div class="phase3-route">' + escapeHtml(phase3RouteText(record)) + '</div>' +
                            phase3ReasonBadges(record) +
                          '</button>';
                        }}).join("");
                      panel.querySelectorAll(".phase3-item").forEach(function (button) {{
                        button.addEventListener("click", function () {{
                          var index = Number(button.getAttribute("data-phase3-index"));
                          focusPhase3Candidate(records[index], config);
                        }});
                      }});
                    }}
                    var boundaryButton = document.getElementById("map-highlight-boundary");
                    if (boundaryButton) {{
                      boundaryButton.addEventListener("click", highlightSelfEnemyFisheryEdges);
                    }}
                    var boundaryDepthButton = document.getElementById("map-highlight-boundary-depth");
                    if (boundaryDepthButton) {{
                      boundaryDepthButton.addEventListener("click", function () {{
                        highlightSelfEnemyFisheryEdgesWithSelfDepth(1);
                      }});
                    }}
                    var boundaryDepth2Button = document.getElementById("map-highlight-boundary-depth-2");
                    if (boundaryDepth2Button) {{
                      boundaryDepth2Button.addEventListener("click", function () {{
                        highlightSelfEnemyFisheryEdgesWithSelfDepth(2);
                      }});
                    }}
                    var boundaryDepth3Button = document.getElementById("map-highlight-boundary-depth-3");
                    if (boundaryDepth3Button) {{
                      boundaryDepth3Button.addEventListener("click", function () {{
                        highlightSelfEnemyFisheryEdgesWithSelfDepth(3);
                      }});
                    }}
                    var enemyThreatButton = document.getElementById("map-highlight-enemy-threat");
                    if (enemyThreatButton) {{
                      enemyThreatButton.addEventListener("click", function () {{
                        highlightCandidateEdges("enemyThreat", "#ef4444", 7);
                      }});
                    }}
                    var pactThreatButton = document.getElementById("map-highlight-pact-threat");
                    if (pactThreatButton) {{
                      pactThreatButton.addEventListener("click", function () {{
                        renderPhase3Candidates("pactThreat");
                      }});
                    }}
                    var friendlyPressureButton = document.getElementById("map-highlight-friendly-pressure");
                    if (friendlyPressureButton) {{
                      friendlyPressureButton.addEventListener("click", function () {{
                        highlightCandidateEdges("friendlyPressure", "#f97316", 7);
                      }});
                    }}
                    var interdictionButton = document.getElementById("map-highlight-interdiction");
                    if (interdictionButton) {{
                      interdictionButton.addEventListener("click", function () {{
                        highlightCandidateEdges("interdiction", "#0ea5e9", 7);
                      }});
                    }}
                    var riskAvoidanceButton = document.getElementById("map-highlight-risk-avoidance");
                    if (riskAvoidanceButton) {{
                      riskAvoidanceButton.addEventListener("click", function () {{
                        highlightCandidateEdges("riskAvoidance", "#be123c", 7);
                      }});
                    }}
                    var phase3DefenseButton = document.getElementById("map-phase3-defense");
                    if (phase3DefenseButton) {{
                      phase3DefenseButton.addEventListener("click", function () {{
                        renderPhase3Candidates("defense");
                      }});
                    }}
                    var phase3AttackButton = document.getElementById("map-phase3-attack");
                    if (phase3AttackButton) {{
                      phase3AttackButton.addEventListener("click", function () {{
                        renderPhase3Candidates("attack");
                      }});
                    }}
                    var phase3InterdictionButton = document.getElementById("map-phase3-interdiction");
                    if (phase3InterdictionButton) {{
                      phase3InterdictionButton.addEventListener("click", function () {{
                        renderPhase3Candidates("interdiction");
                      }});
                    }}
                    var phase3RiskButton = document.getElementById("map-phase3-risk");
                    if (phase3RiskButton) {{
                      phase3RiskButton.addEventListener("click", function () {{
                        renderPhase3Candidates("risk");
                      }});
                    }}
                    var phase3ProtectionButton = document.getElementById("map-phase3-protection");
                    if (phase3ProtectionButton) {{
                      phase3ProtectionButton.addEventListener("click", function () {{
                        renderPhase3Candidates("protection");
                      }});
                    }}
                    var phase3TimeButton = document.getElementById("map-phase3-time");
                    if (phase3TimeButton) {{
                      phase3TimeButton.addEventListener("click", function () {{
                        renderPhase3Candidates("time");
                      }});
                    }}
                    var phase3OverallButton = document.getElementById("map-phase3-overall");
                    if (phase3OverallButton) {{
                      phase3OverallButton.addEventListener("click", function () {{
                        renderPhase3Candidates("overall");
                      }});
                    }}
                    var enemyExpandButton = document.getElementById("map-highlight-enemy-expand");
                    if (enemyExpandButton) {{
                      enemyExpandButton.addEventListener("click", function () {{
                        highlightCandidateEdges("enemyExpand", "#eab308", 6);
                      }});
                    }}
                    var friendlyExpandButton = document.getElementById("map-highlight-friendly-expand");
                    if (friendlyExpandButton) {{
                      friendlyExpandButton.addEventListener("click", function () {{
                        highlightCandidateEdges("friendlyExpand", "#22c55e", 6);
                      }});
                    }}
                    var clearEdgeButton = document.getElementById("map-clear-edge-highlight");
                    if (clearEdgeButton) {{
                      clearEdgeButton.addEventListener("click", clearAllEdgeHighlights);
                    }}
                    var routeIgnoreButton = document.getElementById("map-route-ignore-power");
                    if (routeIgnoreButton) {{
                      routeIgnoreButton.addEventListener("click", function () {{
                        startRouteMode("ignore");
                      }});
                    }}
                    var routeWeakerButton = document.getElementById("map-route-weaker-power");
                    if (routeWeakerButton) {{
                      routeWeakerButton.addEventListener("click", function () {{
                        startRouteMode("weaker");
                      }});
                    }}
                    var clearRouteButton = document.getElementById("map-clear-route");
                    if (clearRouteButton) {{
                      clearRouteButton.addEventListener("click", clearAllEdgeHighlights);
                    }}
                    async function postRefresh(endpoint) {{
                      var response = await fetch(endpoint, {{
                        method: "POST",
                        headers: {{ "Content-Type": "application/json" }}
                      }});
                      var text = await response.text();
                      var data = {{}};
                      if (text) {{
                        try {{
                          data = JSON.parse(text);
                        }} catch (error) {{
                          data = {{ error: text }};
                        }}
                      }}
                      if (!response.ok || data.ok === false) {{
                        throw new Error(data.stderr || data.error || (response.status + " " + response.statusText));
                      }}
                      return data;
                    }}
                    async function refreshMapFromSheet() {{
                      var refreshButton = document.getElementById("map-refresh-sheet");
                      if (!refreshButton) return;
                      var originalText = refreshButton.textContent;
                      refreshButton.disabled = true;
                      refreshButton.textContent = "\\u66f4\\u65b0\\u4e2d...";
                      var endpoints = ["/api/refresh"];
                      var host = window.location.hostname;
                      if ((host === "127.0.0.1" || host === "localhost") && window.location.port !== "8010") {{
                        endpoints.push(window.location.protocol + "//" + host + ":8010/api/refresh");
                      }}
                      var lastError = null;
                      for (var i = 0; i < endpoints.length; i += 1) {{
                        try {{
                          await postRefresh(endpoints[i]);
                          window.location.href = window.location.pathname + "?updated=" + Date.now();
                          return;
                        }} catch (error) {{
                          lastError = error;
                        }}
                      }}
                      refreshButton.disabled = false;
                      refreshButton.textContent = originalText;
                      alert(
                        "\\u30de\\u30c3\\u30d7\\u3092\\u66f4\\u65b0\\u3067\\u304d\\u307e\\u305b\\u3093\\u3067\\u3057\\u305f\\u3002\\n" +
                        "\\u5148\\u306b interactive_server.py \\u3092 --port 8010 \\u3067\\u8d77\\u52d5\\u3057\\u3066\\u304f\\u3060\\u3055\\u3044\\u3002\\n\\n" +
                        (lastError ? lastError.message : "")
                      );
                    }}
                    var refreshButton = document.getElementById("map-refresh-sheet");
                    if (refreshButton) {{
                      refreshButton.addEventListener("click", refreshMapFromSheet);
                    }}
                    window.highlightConnectedEdges = highlightConnectedEdges;
                    window.highlightSelfEnemyFisheryEdges = highlightSelfEnemyFisheryEdges;
                    window.highlightSelfEnemyFisheryEdgesWithSelfDepth = highlightSelfEnemyFisheryEdgesWithSelfDepth;
                    window.highlightCandidateEdges = highlightCandidateEdges;
                    window.resetEdgeHighlights = resetEdgeHighlights;
                    window.showShortestRoute = showShortestRoute;
                    window.refreshMapFromSheet = refreshMapFromSheet;
                    function renderNodeInfo(nodeId) {{
                      var panel = document.getElementById("node-info-panel");
                      if (!panel || !nodeId || !nodes) return;
                      var node = nodes.get(nodeId);
                      if (!node) return;
                      var memo = valueOrDash(node.memo);
                      panel.innerHTML =
                        "<h2>" + escapeHtml(valueOrDash(node.area) + " " + valueOrDash(node.coord)) + "</h2>" +
                        "<dl>" +
                        row("Position key", node.id) +
                        row("Type", node.nodeType) +
                        row("Alliance", node.displayOwner || node.ownerLabel || node.owner) +
                        row("Alliance power", node.alliancePowerLabel || "-") +
                        row("Alliance rank", node.alliancePowerRank ? ("#" + node.alliancePowerRank + " / " + valueOrDash(node.alliancePowerServer)) : "-") +
                        row("Alliance name", node.allianceName || "-") +
                        row("Status", node.mapStatus) +
                        row("Acquired", node.acquiredAt) +
                        row("Protect until", node.rawProtectUntil || node.protectUntil) +
                        row("Hours left", node.hoursRemaining === null ? "-" : node.hoursRemaining + " h") +
                        row("Importance", node.importance) +
                        (memo === "-" ? "" : "<dd class=\\"node-info-memo\\">" + escapeHtml(memo) + "</dd>") +
                        "</dl>";
                    }}
                    function resetNodeInfo() {{
                      var panel = document.getElementById("node-info-panel");
                      if (panel) {{
                        panel.innerHTML = '<div class="node-info-empty">' + escapeHtml(emptyText) + '</div>';
                      }}
                    }}
                    window.renderNodeInfo = renderNodeInfo;
                    network.on("selectNode", function (params) {{
                      var nodeId = params.nodes && params.nodes[0];
                      renderNodeInfo(nodeId);
                      if (!handleRouteSelection(nodeId)) {{
                        highlightConnectedEdges(nodeId);
                      }}
                    }});
                    network.on("deselectNode", function () {{
                      resetNodeInfo();
                    }});
                  }})();
"""
    if "</head>" in html:
        html = html.replace("</head>", style + "\n</head>", 1)
    if "<body>" in html:
        html = html.replace("<body>", "<body>\n" + panel, 1)
    return html.replace("                  return network;", attach_js + "\n                  return network;", 1)


def clean_generated_html(html: str) -> str:
    return "\n".join(line.rstrip() for line in html.splitlines()) + "\n"


def build_state_payload(
    graph: nx.Graph,
    edges: list[Edge],
    analysis: dict[str, Any],
    now: datetime,
    config: dict[str, Any],
    tz: ZoneInfo,
    alliance_powers: dict[str, AlliancePower],
    simulation_config: dict[str, Any] | None = None,
    pacts: dict[str, Any] | None = None,
) -> dict[str, Any]:
    warning_hours = float(config.get("protect_warning_hours", 6))
    owner_affiliations = build_owner_affiliations(graph, config)
    nodes = []
    for _, node_data in graph.nodes(data=True):
        node = Node(**{field: node_data[field] for field in Node.__dataclass_fields__})
        record = asdict(node)
        visual_x, visual_y = visual_position(node, config)
        record["visual_x"] = visual_x
        record["visual_y"] = visual_y
        record["display_owner"] = display_node_owner(node)
        record["destroyed"] = is_destroyed_node(node)
        record["affiliation"] = strategic_affiliation(node, owner_affiliations)
        record["strategic_color"] = strategic_color(node, owner_affiliations)
        record["protection"] = protection_status(node, now, warning_hours, tz)
        record["alliance_power"] = alliance_power_payload(node.owner, alliance_powers)
        nodes.append(record)
    alliance_rankings = sorted(
        {item.alliance: item for item in alliance_powers.values()}.values(),
        key=lambda item: (item.overall_rank is None, item.overall_rank or 999999, item.alliance),
    )
    payload = {
        "generated_at": now.isoformat(),
        "timezone": str(tz),
        "nodes": nodes,
        "connections": [asdict(edge) for edge in edges],
        "alliance_power_rankings": [alliance_power_to_record(item) for item in alliance_rankings],
        "owner_affiliations": owner_affiliations,
        "pacts": pacts or {"active": [], "partners_by_alliance": {}, "active_count": 0},
        **analysis,
    }
    payload["invasion_simulation"] = build_state_invasion_simulation(payload, simulation_config or {})
    return payload


def write_json_payload(payload: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_json(
    graph: nx.Graph,
    edges: list[Edge],
    analysis: dict[str, Any],
    output_path: Path,
    now: datetime,
    config: dict[str, Any],
    tz: ZoneInfo,
    alliance_powers: dict[str, AlliancePower],
    simulation_config: dict[str, Any] | None = None,
    pacts: dict[str, Any] | None = None,
) -> None:
    payload = build_state_payload(graph, edges, analysis, now, config, tz, alliance_powers, simulation_config, pacts)
    write_json_payload(payload, output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build LastWar invasion strategy graph outputs.")
    parser.add_argument("--config", help="JSON config path.")
    parser.add_argument("--nodes", help="Override node CSV path.")
    parser.add_argument("--edges", help="Override edge CSV path.")
    parser.add_argument("--html", help="Override HTML output path.")
    parser.add_argument("--json", dest="json_output", help="Override JSON output path.")
    parser.add_argument("--shortest-from", help="Shortest-path source node id.")
    parser.add_argument("--shortest-to", help="Shortest-path target node id.")
    parser.add_argument("--now", help="Override current time for repeatable tests. ISO-8601.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path.cwd()
    config = read_config(args.config)
    tz = ZoneInfo(config.get("timezone", "Asia/Tokyo"))
    alliance_powers = load_alliance_powers(config.get("alliance_power", {}), repo_root)

    sources = config.get("sources", {})
    if args.nodes:
        sources["nodes"] = {"type": "csv", "path": args.nodes}
    if args.edges:
        sources["edges"] = {"type": "csv", "path": args.edges}
    if "nodes" not in sources:
        raise ValueError("Node source is required")

    node_rows = read_csv_rows(sources["nodes"], repo_root)
    source_config = config.get("source_options", {})
    nodes, adjacency_edges = load_nodes(node_rows, source_config)
    explicit_edges = []
    if sources.get("edges"):
        explicit_edges = load_edges(read_csv_rows(sources["edges"], repo_root))
    edge_config = config.get("edge_derivation", {})
    derived_edges = derive_distance_edges(nodes, edge_config)
    edges = filter_edges_by_type_rules(dedupe_edges(adjacency_edges + explicit_edges + derived_edges), nodes, edge_config)

    graph = build_graph(nodes, edges)
    shortest = config.get("shortest_path", {})
    shortest_from = args.shortest_from or shortest.get("from")
    shortest_to = args.shortest_to or shortest.get("to")
    analysis_config = config.get("analysis", {})
    analysis = analyze_graph(graph, analysis_config, shortest_from, shortest_to)
    analysis["alliance_power_source"] = config.get("alliance_power", {})

    now = parse_optional_time(args.now, tz) if args.now else datetime.now(tz)
    assert now is not None
    pact_payload = {"active": [], "partners_by_alliance": {}, "active_count": 0}
    if sources.get("pacts"):
        pact_payload = load_pacts(read_csv_rows(sources["pacts"], repo_root), now, tz, sources["pacts"])

    output_config = config.get("output", {})
    html_path = Path(args.html or output_config.get("html", "sample_output/map.html"))
    json_path = Path(args.json_output or output_config.get("json", "sample_output/state.json"))
    briefing_path = Path(output_config.get("briefing", "sample_output/briefing_input.json"))
    state_payload = build_state_payload(graph, edges, analysis, now, analysis_config, tz, alliance_powers, config.get("simulation", {}), pact_payload)
    write_html(graph, analysis, repo_root / html_path, now, analysis_config, tz, alliance_powers, state_payload.get("invasion_simulation"))
    write_json_payload(state_payload, repo_root / json_path)
    write_json_payload(build_briefing_input(state_payload, state_payload.get("invasion_simulation"), config.get("simulation", {})), repo_root / briefing_path)
    print(f"Wrote {html_path}")
    print(f"Wrote {json_path}")
    print(f"Wrote {briefing_path}")
    print(f"Critical nodes: {len(analysis['critical_nodes'])}")
    return 0


def dedupe_edges(edges: list[Edge]) -> list[Edge]:
    seen: set[tuple[str, str]] = set()
    deduped: list[Edge] = []
    for edge in edges:
        key = tuple(sorted((edge.source, edge.target)))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(edge)
    return deduped


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
