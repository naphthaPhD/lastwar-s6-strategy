from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_EXCEL = "s6powerrank_8server_power_2026-05-10_en.xlsx"
DEFAULT_SHEET = "Alliance Power"
DEFAULT_UNMATCHED_OWNERS = "sample_output/sheet_migration/unmatched_owners.csv"
DEFAULT_ALLIANCE_DIRECTORY = "sample_output/sheet_migration/alliance_directory.csv"
DEFAULT_PACTS = "sample_output/sheet_migration/pacts.csv"
DEFAULT_CANDIDATES_OUTPUT = "sample_output/sheet_migration/alliance_directory_from_power_candidates.csv"
DEFAULT_TOP_OUTPUT = "sample_output/sheet_migration/alliance_directory_from_power_top.csv"
SOURCE = "alliance_power_2026-05-10"
TOP_LIMIT = 30


@dataclass(frozen=True)
class PowerAlliance:
    alliance: str
    server: str
    side: str
    power: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Suggest alliance_directory.csv additions from an Alliance Power workbook."
    )
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Input power ranking .xlsx path.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Worksheet name containing Alliance Power.")
    parser.add_argument(
        "--unmatched-owners",
        default=DEFAULT_UNMATCHED_OWNERS,
        help="Input unmatched_owners.csv path.",
    )
    parser.add_argument(
        "--alliance-directory",
        default=DEFAULT_ALLIANCE_DIRECTORY,
        help="Existing alliance_directory.csv path. Read only.",
    )
    parser.add_argument("--pacts", default=DEFAULT_PACTS, help="Existing pacts.csv path. Read only.")
    parser.add_argument(
        "--candidates-output",
        default=DEFAULT_CANDIDATES_OUTPUT,
        help="Output alliance_directory_from_power_candidates.csv path.",
    )
    parser.add_argument(
        "--top-output",
        default=DEFAULT_TOP_OUTPUT,
        help="Output alliance_directory_from_power_top.csv path.",
    )
    return parser.parse_args()


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return [{key: (value or "").strip() for key, value in row.items()} for row in csv.DictReader(fh)]


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def normalize_server(value: Any) -> str:
    return str(value or "").strip().lstrip("#")


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def side_to_affiliation(side: str) -> str:
    normalized = str(side or "").strip().lower()
    if normalized in {"ally", "friend", "friendly"}:
        return "ally"
    if normalized == "enemy":
        return "enemy"
    return "unknown"


def pact_server_rules(path: Path) -> dict[str, tuple[str, str]]:
    rules: dict[str, tuple[str, str]] = {}
    for row in read_csv(path):
        server = normalize_server(row.get("server", ""))
        alliance = row.get("alliance", "")
        if not server or alliance:
            continue
        rules.setdefault(
            server,
            (
                row.get("affiliation", "unknown") or "unknown",
                row.get("pact_status", "unknown") or "unknown",
            ),
        )
    return rules


def directory_keys(path: Path) -> set[str]:
    keys: set[str] = set()
    for row in read_csv(path):
        alliance = row.get("alliance", "")
        if alliance:
            keys.add(alliance)
        for alias in row.get("alliance_alias", "").split(","):
            alias = alias.strip()
            if alias:
                keys.add(alias)
    return keys


def unmatched_owners(path: Path) -> list[dict[str, str]]:
    rows = []
    for row in read_csv(path):
        owner_alliance = row.get("owner_alliance", "")
        if not owner_alliance:
            continue
        rows.append(row)
    return rows


def find_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    required = {"server", "alliance", "side", "power"}
    for index, row in enumerate(rows):
        header = {normalize_header(value): col_index for col_index, value in enumerate(row)}
        if required.issubset(header):
            return index, header
    raise ValueError("Could not find Alliance Power header row with Server, Alliance, Side, and Power.")


def read_power_alliances(excel_path: Path, sheet_name: str) -> dict[str, PowerAlliance]:
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {sheet_name}")
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    header_index, header = find_header(rows)

    alliances: dict[str, PowerAlliance] = {}
    for row in rows[header_index + 1 :]:
        alliance = str(row[header["alliance"]] or "").strip()
        if not alliance or alliance in alliances:
            continue
        server = normalize_server(row[header["server"]])
        side = str(row[header["side"]] or "").strip()
        power_value = row[header["power"]]
        power = "" if power_value is None else str(power_value)
        alliances[alliance] = PowerAlliance(
            alliance=alliance,
            server=server,
            side=side,
            power=power,
        )
    return alliances


def suggestion_from_power(power: PowerAlliance, pact_rules: dict[str, tuple[str, str]]) -> tuple[str, str, str]:
    if power.server in pact_rules:
        affiliation, pact_status = pact_rules[power.server]
        confidence = "high" if affiliation != "unknown" else "unknown"
        return affiliation, pact_status, confidence

    affiliation = side_to_affiliation(power.side)
    if affiliation == "unknown":
        return "unknown", "unknown", "unknown"
    return affiliation, "attack_allowed", "medium"


def build_candidate_rows(
    unmatched_rows: list[dict[str, str]],
    power_alliances: dict[str, PowerAlliance],
    directory_key_set: set[str],
    pact_rules: dict[str, tuple[str, str]],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    candidate_rows: list[dict[str, str]] = []
    missing_rows: list[dict[str, str]] = []

    for row in unmatched_rows:
        owner_alliance = row.get("owner_alliance", "")
        node_count = row.get("node_count", "0")
        power = power_alliances.get(owner_alliance)
        if not power:
            missing_rows.append(row)
            continue

        affiliation, pact_status, confidence = suggestion_from_power(power, pact_rules)
        already_in_directory = owner_alliance in directory_key_set
        candidate_rows.append(
            {
                "owner_alliance": owner_alliance,
                "matched_power_alliance": power.alliance,
                "server": power.server,
                "side": power.side,
                "suggested_affiliation": affiliation,
                "suggested_pact_status": pact_status,
                "power": power.power,
                "node_count": node_count,
                "source": SOURCE,
                "confidence": confidence,
                "reason": "exact_alliance_match",
                "already_in_directory": "TRUE" if already_in_directory else "FALSE",
                "note": "dictionary candidate only; not ownership evidence",
            }
        )

    candidate_rows.sort(key=lambda item: (-int(item["node_count"] or 0), item["owner_alliance"]))
    missing_rows.sort(key=lambda item: (-int(item.get("node_count", "0") or 0), item.get("owner_alliance", "")))
    return candidate_rows, missing_rows


def print_inspection(
    unmatched_rows: list[dict[str, str]],
    power_alliances: dict[str, PowerAlliance],
    candidate_rows: list[dict[str, str]],
    missing_rows: list[dict[str, str]],
) -> None:
    matched_node_count = sum(int(row["node_count"] or 0) for row in candidate_rows)
    already_count = sum(1 for row in candidate_rows if row["already_in_directory"] == "TRUE")
    add_count = sum(1 for row in candidate_rows if row["already_in_directory"] == "FALSE")

    print("Alliance directory from power inspection")
    print(f"unmatched_owner_alliance_count={len(unmatched_rows)}")
    print(f"power_alliance_unique_count={len(power_alliances)}")
    print(f"matched_count={len(candidate_rows)}")
    print(f"unmatched_in_power_count={len(missing_rows)}")
    print(f"matched_node_count={matched_node_count}")
    print(f"already_in_directory_count={already_count}")
    print(f"candidate_to_add_count={add_count}")
    print("top_candidates")
    for row in candidate_rows[:20]:
        print(
            f"{row['owner_alliance']}\t"
            f"nodes={row['node_count']}\t"
            f"server={row['server']}\t"
            f"side={row['side']}\t"
            f"affiliation={row['suggested_affiliation']}\t"
            f"already_in_directory={row['already_in_directory']}"
        )
    if missing_rows:
        print("unmatched_in_power")
        print(",".join(row["owner_alliance"] for row in missing_rows))


def main() -> int:
    args = parse_args()
    excel_path = Path(args.excel)
    unmatched_path = Path(args.unmatched_owners)
    directory_path = Path(args.alliance_directory)
    pacts_path = Path(args.pacts)
    candidates_output_path = Path(args.candidates_output)
    top_output_path = Path(args.top_output)

    power_alliances = read_power_alliances(excel_path, args.sheet)
    unmatched_rows = unmatched_owners(unmatched_path)
    candidate_rows, missing_rows = build_candidate_rows(
        unmatched_rows,
        power_alliances,
        directory_keys(directory_path),
        pact_server_rules(pacts_path),
    )

    fieldnames = [
        "owner_alliance",
        "matched_power_alliance",
        "server",
        "side",
        "suggested_affiliation",
        "suggested_pact_status",
        "power",
        "node_count",
        "source",
        "confidence",
        "reason",
        "already_in_directory",
        "note",
    ]
    write_csv(candidates_output_path, candidate_rows, fieldnames)
    write_csv(top_output_path, candidate_rows[:TOP_LIMIT], fieldnames)
    print_inspection(unmatched_rows, power_alliances, candidate_rows, missing_rows)
    print(f"candidates_output={candidates_output_path}")
    print(f"top_output={top_output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
